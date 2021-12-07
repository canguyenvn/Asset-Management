
import sys
import shutil
import os
import time
from datetime import date,datetime,timedelta
from PyQt5 import uic
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QLabel, QDialog, QApplication, QMessageBox, QTableView,QShortcut,QCompleter,QCalendarWidget
from PyQt5.QtGui import QFont, QIcon, QKeySequence, QRegExpValidator
import PyQt5.QtCore
from PyQt5.QtCore import QAbstractTableModel, Qt, QTimer, QDateTime, QRegExp
import pandas as pd
import sqlite3
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
import subprocess
import socket
import resfolder

class ShareTheSameMethod():
    def GetUserName(self):
        try:
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            query=f"""SELECT "Full Name"
                FROM UserHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            lastRow=len(df)
            sqlite_insert_logger_query= f"""SELECT "Full Name" FROM UserHistory WHERE rowid = ?"""
            data_tuple = (lastRow,)
            cursor.execute(sqlite_insert_logger_query, data_tuple)
            Sender=cursor.fetchone()[0]
            sqliteConnection.commit()
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 
        return Sender

    def BackToDashBoard(self):
        try:
            photoInLinePath
            photoPath
            if os.path.exists(photoInLinePath) and os.path.exists(photoPath):
                os.remove(photoInLinePath)
                os.remove(photoPath)
        except:
            pass
        DB=DashBoard()
        widget.addWidget(DB)
        widget.setFixedHeight(604)
        widget.setFixedWidth(800)
        widget.setWindowTitle("DashBoard Window")
        widget.setWindowIcon(QIcon(":icon/DashBoard.png"))
        widget.currentWidget().deleteLater()
        widget.setCurrentWidget(DB)
        widget.move(screen_center(widget))

    def ExitSystemQuick(self):
        widget.deleteLater()
        try:
            photoInLinePath
            photoPath
            if os.path.exists(photoInLinePath) and os.path.exists(photoPath):
                os.remove(photoInLinePath)
                os.remove(photoPath)
        except:
            pass
        #xóa database tạm của Inventory Window
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor      = sqliteConnection.cursor()
        try:
            dropInventoryTable = "DROP TABLE IF EXISTS Inventory"
            cursor.execute(dropInventoryTable)
            dropsearchResultTable = "DROP TABLE IF EXISTS SearchResult"
            cursor.execute(dropsearchResultTable)
            dropsearchResultTable = "DROP TABLE IF EXISTS Filter"
            cursor.execute(dropsearchResultTable)
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 
        #cập nhật file log thời gian đăng xuất
        try:
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            query=f"""SELECT Session
            FROM UserHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            lastRow=len(df)
            if lastRow==0:
                return
            if df.loc[lastRow-1].iat[0]=="":
                noOperation='This user read only, no operation changes the database.'
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                SET "Logout Time"=?,
                                    Session=?
                                WHERE rowid = ?"""
                data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),noOperation,lastRow)
            else:
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                SET "Logout Time"=?
                                WHERE rowid = ?"""
                data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),lastRow)
            cursor.execute(sqlite_insert_logger_query, data_tuple)
            sqliteConnection.commit()
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 
        
    def ExitSystem(self):
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("ShareTheSameMethod", "確認"),QtWidgets.QApplication.translate("ShareTheSameMethod", "システムを終了してもよろしいですか？"))
        if respond==QMessageBox.Ok:
            widget.deleteLater()
            try:
                photoInLinePath
                photoPath
                if os.path.exists(photoInLinePath) and os.path.exists(photoPath):
                    os.remove(photoInLinePath)
                    os.remove(photoPath)
            except:
                pass
            #xóa database tạm của Inventory Window
            try:
                sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
                cursor      = sqliteConnection.cursor()
                dropInventoryTable = "DROP TABLE IF EXISTS Inventory"
                cursor.execute(dropInventoryTable)
                dropsearchResultTable = "DROP TABLE IF EXISTS SearchResult"
                cursor.execute(dropsearchResultTable)
                dropsearchResultTable = "DROP TABLE IF EXISTS Filter"
                cursor.execute(dropsearchResultTable)
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 
            #cập nhật file log thời gian đăng xuất
            try:
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT Session
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                if lastRow==0:
                    return
                if df.loc[lastRow-1].iat[0]=="":
                    noOperation='This user read only, the database is not changed.'
                    sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET "Logout Time"=?,
                                        Session=?
                                    WHERE rowid = ?"""
                    data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),noOperation,lastRow)
                else:
                    sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET "Logout Time"=?
                                    WHERE rowid = ?"""
                    data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 
            #widget.deleteLater()

    def MessageBoxOK(self,icon,flat,string):
        msg = QMessageBox()
        if icon=="Information":
            msg.setIcon(QMessageBox.Information)
        elif icon=="Warning":
            msg.setIcon(QMessageBox.Warning)
        elif icon=="Critical":
            msg.setIcon(QMessageBox.Critical)
        font = QFont()
        font.setFamily("Yu Mincho")
        font.setPointSize(13)
        msg.setFont(font)
        msg.setText(string)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setWindowTitle(flat)
        msg.exec_()

    def MessageBoxOKCancel(self,flat,string):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Question)
        font = QFont()
        font.setFamily("Yu Mincho")
        font.setPointSize(13)
        msg.setFont(font)
        msg.setText(string)
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msg.setWindowTitle(flat)
        return msg.exec_()

class MaintenanceHistoryTableModel(PyQt5.QtCore.QAbstractTableModel,ShareTheSameMethod):

    def __init__(self, data):
        super(MaintenanceHistoryTableModel, self).__init__()
        self._data = data

    def data(self, index, role):
        if role == Qt.DisplayRole:
            value = self._data.iloc[index.row(), index.column()]
            return str(value)

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

    def headerData(self, section, orientation, role):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._data.columns[section]

            if orientation == Qt.Vertical:
                return self._data.index[section]+1
    
    def flags(self, index):
        ColEnable=[2,3,4,5,6]
        if index.column() in ColEnable:
            return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable
        else:
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable

    def setData(self, index, value, role):
        if role==Qt.EditRole:
            respond=self.MessageBoxOKCancel("確認","本気ですか？")
            if respond==QMessageBox.Ok:
                self._data.iloc[index.row(),index.column()]=value
                key=Eq.lineEdit_serial.text()
                try:
                    sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
                    cursor = sqliteConnection.cursor()
                    if index.column()==2:
                        sqlite_update_record_query= f"""UPDATE MaintenanceHistory 
                                    SET 日付=?
                                    WHERE 製造番号=?"""
                    elif index.column()==3:
                        sqlite_update_record_query= f"""UPDATE MaintenanceHistory 
                                    SET 箇所=?
                                    WHERE 製造番号=?"""
                    elif index.column()==4:
                        sqlite_update_record_query= f"""UPDATE MaintenanceHistory 
                                    SET 実施内容=?
                                    WHERE 製造番号=?"""
                    elif index.column()==5:
                        sqlite_update_record_query= f"""UPDATE MaintenanceHistory 
                                    SET 担当者=?
                                    WHERE 製造番号=?"""
                    elif index.column()==6:
                        sqlite_update_record_query= f"""UPDATE MaintenanceHistory 
                                    SET 確認者=?
                                    WHERE 製造番号=?"""
                    data_tuple = (value,key)
                    cursor.execute(sqlite_update_record_query, data_tuple)
                    sqliteConnection.commit()
                    print("Cập nhật database thành công")
                    cursor.close()
                    self.MessageBoxOK("確認","更新成功！")
                    PyQt5.QtCore.QTimer.singleShot(0, Eq.SearchText.setFocus)
                    Eq.UpdateSearchUI()
                    Eq.ResetFieldForm()
                except sqlite3.Error as error:
                    print("Lỗi", error)
                finally:
                    if sqliteConnection:
                        sqliteConnection.close()
                return True
            else:
                return False

class PandasModelForFilter(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, parnet=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            value = self._data.iloc[index.row(), index.column()]
            if role == Qt.DisplayRole or role == Qt.EditRole:
                return  str(value)
            if role == Qt.BackgroundRole and value == 'Yes':
                return PyQt5.QtGui.QColor('red')
            if role == Qt.DecorationRole:
                if value=='':
                    return PyQt5.QtGui.QIcon(':img/edit.png')
            
    def headerData(self, col, orientation, role):
        if role==Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._data.columns[col]
            if orientation==Qt.Vertical:
                return self._data.index[col]+1

class PandasModelForImportHistoryAndSearch(QAbstractTableModel,ShareTheSameMethod):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, parnet=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            value = self._data.iloc[index.row(), index.column()]
            if role == Qt.DisplayRole or role == Qt.EditRole:
                return  str(value)

    def headerData(self, col, orientation, role):
        if role==Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._data.columns[col]
            if orientation==Qt.Vertical:
                return self._data.index[col]+1

class PandasModelForExportHistory(QAbstractTableModel,ShareTheSameMethod):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, parnet=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            value = self._data.iloc[index.row(), index.column()]
            if role == Qt.DisplayRole or role == Qt.EditRole:
                return  str(value)
            if role == Qt.DecorationRole:
                if value=='':
                    return PyQt5.QtGui.QIcon(':img/edit.png')
            
    def headerData(self, col, orientation, role):
        if role==Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._data.columns[col]
            if orientation==Qt.Vertical:
                return self._data.index[col]+1

class Logger(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(Logger,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/LoggerWindow.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        self.InitialValue()
        self.Signal_Slot()

    def InitialValue(self):
        self.label.setText(QtWidgets.QApplication.translate("Logger", "システムにログインするユーザー情報"))
        self.ExitButton.setText(QtWidgets.QApplication.translate("Logger", "終了"))
        self.ReturnButton.setText(QtWidgets.QApplication.translate("Logger", "戻る"))

        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # update sau mỗi giây
        self.showTime()

    def Signal_Slot(self):
        self.ReturnButton.clicked.connect(self.BackToDashBoard)
        self.ExitButton.clicked.connect(self.ExitSystem)
        self.SaveExcelButton.clicked.connect(self.SaveExcelLog)
        self.shortcut_toDashBoardWindow = QShortcut(QKeySequence('Ctrl+B'), self)
        self.shortcut_toDashBoardWindow.activated.connect(self.BackToDashBoard)
        self.shortcut_ExitWindow = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_ExitWindow.activated.connect(self.ExitSystemQuick)

        try:
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            query=f"""SELECT *
            FROM UserHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            if not df.size==0:
                if df.size>200:
                    df=df.tail(100)
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForExportHistory(reversed_df)
                self.tableView_Logger.setModel(self.model)
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()
    
    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_3.setText(displayTxt)

    def __del__(self):
        print("hàm hủy của Logger đã được gọi!")

    def SaveExcelLog(self):
        sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
        try:
            dfInventory = pd.read_sql("SELECT * FROM UserHistory;", con=sqliteConnection)
            if not dfInventory.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Logger", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/UserLoginHistory_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfInventory.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:E'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Logger", "確認"),QtWidgets.QApplication.translate("Logger", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

class Inventory(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(Inventory,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/InventoryWindowTest.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        self.InitialValue()
        self.Signal_Slot()

    def __del__(self):
        print("hàm hủy của Inventory đã được gọi!")

    def InitialValue(self):
        self.SearchButton.setText(QtWidgets.QApplication.translate("Inventory", "検索"))
        self.SaveExcelInventoryButton.setText(QtWidgets.QApplication.translate("Inventory", "全部品の在庫"))
        self.SaveExcelSearchButton.setText(QtWidgets.QApplication.translate("Inventory", " 検索結果"))
        self.label_4.setText(QtWidgets.QApplication.translate("Inventory", "カメラモジュール設備技術課の部品と設備管理システム"))
        self.SearchText.setPlaceholderText(QtWidgets.QApplication.translate("Inventory", "例え：fil...,speed...,...")) 
        #DỊCH CHO TAB IMPORT
        self.tabWidget.setTabText(0,QtWidgets.QApplication.translate("Inventory", "部品輸入")) 
        self.label_3.setText(QtWidgets.QApplication.translate("Inventory", "日付"))
        self.label_5.setText(QtWidgets.QApplication.translate("Inventory", "部品名"))
        self.label_6.setText(QtWidgets.QApplication.translate("Inventory", "部品コード"))
        self.label_7.setText(QtWidgets.QApplication.translate("Inventory", "供給者"))
        self.label_8.setText(QtWidgets.QApplication.translate("Inventory", "原産地"))
        self.label_9.setText(QtWidgets.QApplication.translate("Inventory", "使用設備"))
        self.label_10.setText(QtWidgets.QApplication.translate("Inventory", "設備機種"))
        self.label_11.setText(QtWidgets.QApplication.translate("Inventory", "数量"))
        self.label_12.setText(QtWidgets.QApplication.translate("Inventory", "部品パラメータ"))

        self.ImportButton.setText(QtWidgets.QApplication.translate("Inventory", "確認"))
        self.DeleteImport.setText(QtWidgets.QApplication.translate("Inventory", "消去"))
        self.EditImport.setText(QtWidgets.QApplication.translate("Inventory", "編集"))
        self.ExitImportButton.setText(QtWidgets.QApplication.translate("Inventory", "終了"))
        self.ResetImportButton.setText(QtWidgets.QApplication.translate("Inventory", "再設定"))
        self.ReturnImportButton.setText(QtWidgets.QApplication.translate("Inventory", "戻る"))
        #DỊCH CHO TAB EXPORT
        self.tabWidget.setTabText(1,QtWidgets.QApplication.translate("Inventory", "部品輸出")) 
        self.label_13.setText(QtWidgets.QApplication.translate("Inventory", "日付"))
        self.label_14.setText(QtWidgets.QApplication.translate("Inventory", "部品名"))
        self.label_15.setText(QtWidgets.QApplication.translate("Inventory", "部品コード"))
        self.label_16.setText(QtWidgets.QApplication.translate("Inventory", "取出人 "))
        self.label_17.setText(QtWidgets.QApplication.translate("Inventory", "受取人"))
        self.label_18.setText(QtWidgets.QApplication.translate("Inventory", "使用設備"))
        self.label_19.setText(QtWidgets.QApplication.translate("Inventory", "設備機種"))
        self.label_20.setText(QtWidgets.QApplication.translate("Inventory", "数量"))
        self.label_21.setText(QtWidgets.QApplication.translate("Inventory", "理由"))

        self.ExportButton.setText(QtWidgets.QApplication.translate("Inventory", "確認"))
        self.DeleteExport.setText(QtWidgets.QApplication.translate("Inventory", "消去"))
        self.EditExport.setText(QtWidgets.QApplication.translate("Inventory", "編集"))
        self.ExitExportButton.setText(QtWidgets.QApplication.translate("Inventory", "終了"))
        self.ResetExportButton.setText(QtWidgets.QApplication.translate("Inventory", "再設定"))
        self.ReturnExportButton.setText(QtWidgets.QApplication.translate("Inventory", "戻る"))
        #DỊCH CHO TAB FILTER
        self.tabWidget.setTabText(2,QtWidgets.QApplication.translate("Inventory", "Filter定期交換部品(1)")) 
        self.tabWidget.setTabText(3,QtWidgets.QApplication.translate("Inventory", "Filter定期交換部品(2)")) 
        self.label.setText(QtWidgets.QApplication.translate("Inventory", "Filter部品定期交換歴史"))
        self.pushButton.setText(QtWidgets.QApplication.translate("Inventory", " 終了"))
        self.ReturnButton_2.setText(QtWidgets.QApplication.translate("Inventory", " 戻る"))
        self.label_26.setText(QtWidgets.QApplication.translate("Inventory", " 定期交換部品見積和制御"))

        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # update sau mỗi giây
        self.showTime()

        if str(LoginSys.comboBox.currentText())=="日本語":
            self.calendarWidget_Start.setLocale(PyQt5.QtCore.QLocale(PyQt5.QtCore.QLocale.Japanese))
            self.calendarWidget_End.setLocale(PyQt5.QtCore.QLocale(PyQt5.QtCore.QLocale.Japanese))
        else:
            self.calendarWidget_Start.setLocale(PyQt5.QtCore.QLocale(PyQt5.QtCore.QLocale.English))
            self.calendarWidget_End.setLocale(PyQt5.QtCore.QLocale(PyQt5.QtCore.QLocale.English))
        self.SenderText.setText(self.GetUserName())
        self.DateImportText.setPlaceholderText("yyyy/MM/dd")
        validator = QRegExpValidator(QRegExp("[0-9-/]{10}"))
        self.DateImportText.setValidator(validator)
        self.DateExportText.setPlaceholderText("yyyy/MM/dd")
        self.DateExportText.setValidator(validator)
        #self.DateImportText.setText(TodayIs) 
        #self.DateExportText.setText(TodayIs) 
        self.SaveExcelSearchButton.setEnabled(False)
        self.SaveExcelSearchButton.hide()
        self.DeleteImport.setEnabled(False)
        self.EditImport.setEnabled(False)
        self.DeleteExport.setEnabled(False)
        self.EditExport.setEnabled(False)
        validatorPressureQual = QRegExpValidator(QRegExp("[0-9]{15}"))
        self.QualImportText.setValidator(validatorPressureQual)
        self.QualExportText.setValidator(validatorPressureQual)
        self.calendarWidget_End.setSelectionMode(QCalendarWidget.NoSelection)
        self.LoadPartList()
        self.LoadPartCodeList()
        self.UpdateImportHistory()
        self.UpdateExportHistory()
        self.UpdateInventory()
        #self.Filter()
        PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
    
    def GetDateStart(self,date):
        self.calendarWidget_End.setSelectionMode(QCalendarWidget.SingleSelection)
        global DateMonthYearStart
        DateMonthYearStart=date.toString("yyyy/MM/dd")

    def GetDateEnd(self,date):
        global DateMonthYearEnd
        DateMonthYearEnd=date.toString("yyyy/MM/dd")
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT 日付,数量 FROM ExportHistory
                        WHERE 部品名="Filter" AND 部品コード="WGW.01.12.150" 
                        AND 日付>=? AND 日付<=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(DateMonthYearStart,DateMonthYearEnd))
            if not df.size ==0:
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForImportHistoryAndSearch(reversed_df)
                self.tableView_150.setModel(self.model)
                self.tableView_150.setColumnWidth(0, 80)
                self.tableView_150.setColumnWidth(1, 60)
                self.tableView_150.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

            query=f"""SELECT 日付,数量 FROM ExportHistory
                        WHERE 部品名="Filter" AND 部品コード="WGW.01.12.250" 
                        AND 日付>=? AND 日付<=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(DateMonthYearStart,DateMonthYearEnd))
            if not df.size ==0:
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForImportHistoryAndSearch(reversed_df)
                self.tableView_250.setModel(self.model)
                self.tableView_250.setColumnWidth(0, 80)
                self.tableView_250.setColumnWidth(1, 60)
                self.tableView_250.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

            query=f"""SELECT 日付,数量 FROM ExportHistory
                        WHERE 部品名="Filter" AND 部品コード="WGW.01.12.251" 
                        AND 日付>=? AND 日付<=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(DateMonthYearStart,DateMonthYearEnd))
            if not df.size ==0:
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForImportHistoryAndSearch(reversed_df)
                self.tableView_251.setModel(self.model)
                self.tableView_251.setColumnWidth(0, 80)
                self.tableView_251.setColumnWidth(1, 60)
                self.tableView_251.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

            query=f"""SELECT 日付,数量 FROM ExportHistory
                        WHERE 部品名="Filter" AND 部品コード="WGW.02.02.152" 
                        AND 日付>=? AND 日付<=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(DateMonthYearStart,DateMonthYearEnd))
            if not df.size ==0:
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForImportHistoryAndSearch(reversed_df)
                self.tableView_152.setModel(self.model)
                self.tableView_152.setColumnWidth(0, 80)
                self.tableView_152.setColumnWidth(1, 60)
                self.tableView_152.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

            sqliteConnection.commit()
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def LoadPartList(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT PartName FROM PartList""" 
            cursor.execute(query)
            sqliteConnection.commit()
            final_result = [i[0] for i in cursor.fetchall()]
            cursor.close()
            completer = QCompleter(final_result)
            self.PartNameImportText.setCompleter(completer)
            self.PartNameExportText.setCompleter(completer)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def LoadPartCodeList(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT partCode FROM PartCodeList""" 
            cursor.execute(query)
            sqliteConnection.commit()
            final_result = [i[0] for i in cursor.fetchall()]
            cursor.close()
            completer = QCompleter(final_result)
            self.PartCodeExportText.setCompleter(completer)
            self.PartCodeImportText.setCompleter(completer)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_27.setText(displayTxt)

    def Signal_Slot(self):    
        self.ImportButton.clicked.connect(self.AdImportHistory)
        self.ExportButton.clicked.connect(self.AdExportHistory)
        self.SaveExcelImportButton.clicked.connect(self.SaveExcelImportHistory)
        self.SaveExcelExportButton.clicked.connect(self.SaveExcelExportHistory)
        self.ExitImportButton.clicked.connect(self.ExitSystem)
        self.ExitExportButton.clicked.connect(self.ExitSystem)
        self.pushButton.clicked.connect(self.ExitSystem)
        self.ResetImportButton.clicked.connect(self.ResetImportFields)
        self.ResetExportButton.clicked.connect(self.ResetExportFields)
        self.pushButton.clicked.connect(self.ResetExportFields)
        self.ReturnImportButton.clicked.connect(self.BackToDashBoard)
        self.ReturnExportButton.clicked.connect(self.BackToDashBoard)
        self.ReturnButton_2.clicked.connect(self.BackToDashBoard)
        self.SearchButton.clicked.connect(self.SearchParts)
        self.SearchText.returnPressed.connect(self.SearchParts)
        self.SaveExcelInventoryButton.clicked.connect(self.SaveExcelInventory)
        self.SaveExcelSearchButton.clicked.connect(self.SaveExcelSearchResult)
        self.pushButton_2.clicked.connect(self.SaveExcelFilter)
        self.tabWidget.currentChanged.connect (self.SelectTab)
        self.tableViewImport.pressed.connect(self.ShowAllInforImport)
        self.tableViewExport.pressed.connect(self.ShowAllInforExport)
        self.DeleteImport.clicked.connect(self.DeleteImportHistory)
        self.EditImport.clicked.connect(self.EditImportHistory)
        self.DeleteExport.clicked.connect(self.DeleteExportHistory)
        self.EditExport.clicked.connect(self.EditExportHistory)
        self.calendarWidget_Start.clicked.connect(self.GetDateStart)
        self.calendarWidget_End.clicked.connect(self.GetDateEnd)
        self.shortcut_exit = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_exit.activated.connect(self.ExitSystemQuick)
        self.shortcut_returnToDashBoard = QShortcut(QKeySequence('Ctrl+B'), self)
        self.shortcut_returnToDashBoard.activated.connect(self.BackToDashBoard)

    def EditImportHistory(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        DateInput=self.DateImportText.text()
        PartName=self.PartNameImportText.text()
        if PartName=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品名を入力してください！"))
            return
        PartCode=self.PartCodeImportText.text()
        if PartCode=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品コードを入力してください！"))
            return
        Supplier=self.SupplierNameText.text()
        From=self.FromText.text()
        EqUse=self.UseForEqImportText.text()
        EqID=self.ModelEqImportText.text()
        Qual=int(self.QualImportText.text())
        if Qual=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "数量を入力してください！"))
            return
        Des=self.PartDesText.text()
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_insert_new_record_query= f"""UPDATE ImportHistory 
                                SET 日付=?,部品名=?,部品コード=?,供給者=?,原産地=?,使用設備=?,設備機種=?,数量=?,部品パラメータ=?
                                WHERE rowid = ?"""
                data_tuple = (DateInput,PartName,PartCode,Supplier,From,EqUse,EqID,Qual,Des,RowID)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.UpdateImportHistory()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "部品輸入編集成功！"))
            except sqlite3.Error as error:
                print("Lỗi ", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            #cập nhật tên linh kiện mới và mã linh kiện mới vào database
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT partName FROM PartList""" 
                cursor.execute(query)
                final_result = [i[0] for i in cursor.fetchall()]
                if not PartName in final_result:
                    query=f"""INSERT INTO PartList VALUES (?)"""
                    cursor.execute(query,(PartName,))
                
                query=f"""SELECT partCode FROM PartCodeList""" 
                cursor.execute(query)
                final_result = [i[0] for i in cursor.fetchall()]
                if not PartName in final_result:
                    query=f"""INSERT INTO PartCodeList VALUES (?)"""
                    cursor.execute(query,(PartCode,))
                sqliteConnection.commit()
                cursor.close()	
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
            #tính tồn kho cập nhật vô QTableView Search
            self.UpdateInventory()
            self.ResetImportFields()

    def DeleteImportHistory(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_delete_record_query= f"""DELETE from ImportHistory where rowid=?"""
                data_tuple = (RowID,)
                cursor.execute(sqlite_delete_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                #cập nhật logger thông tin xóa form thiết bị
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                PartName=self.PartNameImportText.text()
                PartCode=self.PartCodeImportText.text()
                Quantity=self.QualImportText.text()
                newOperate='This user deleted the part named '+PartName+',code '+PartCode+' ,quantity '+Quantity+' pcs.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.UpdateImportHistory()
                self.UpdateInventory()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "消去成功！"))
                self.ResetImportFields()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
 
    def EditExportHistory(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        DateInput=self.DateExportText.text()
        PartName=self.PartNameExportText.text()
        if PartName=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品名を入力してください！"))
            return
        PartCode=self.PartCodeExportText.text()
        if PartCode=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品コードを入力してください！"))
            return
        Exporter=self.SenderText.text()
        Receiver=self.ReceiverText.text()
        EqUse=self.UseForEqExportText.text()
        EqID=self.ModelEqExportText.text()
        Reason=self.WhyExportText.text()
        Qual=self.QualExportText.text()
        if Qual=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "数量を入力してください！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_insert_new_record_query= f"""UPDATE ExportHistory 
                                SET 日付=?,部品名=?,部品コード=?,取出人=?,受取人=?,使用設備=?,設備機種=?,
                                    数量=?,理由=?
                                WHERE rowid = ?""" 
                data_tuple = (DateInput,PartName,PartCode,Exporter,Receiver,EqUse,EqID,Qual,Reason,RowIDEx)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.UpdateExportHistory()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "部品輸出編集成功！"))
            except sqlite3.Error as error:
                print("Lỗi ", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            #tính tồn kho cập nhật vô QTableView Search
            self.UpdateInventory()
            self.ResetExportFields()

    def DeleteExportHistory(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_delete_record_query= f"""DELETE from ExportHistory where rowid=?"""
                data_tuple = (RowIDEx,)
                cursor.execute(sqlite_delete_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "消去成功！"))
                self.ResetExportFields()
                self.UpdateExportHistory()
                self.UpdateInventory()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()

            #cập nhật logger thông tin xóa form thiết bị
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                PartName=self.PartNameExportText.text()
                PartCode=self.PartCodeExportText.text()
                Quantity=self.QualExportText.text()
                newOperate='This user deleted the part named '+PartName+',code '+PartCode+' ,quantity '+Quantity+' pcs.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()

    def ShowAllInforImport(self,index):
        self.DeleteImport.setEnabled(True)
        self.EditImport.setEnabled(True)
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            findLastRowQuery=f"""SELECT * FROM ImportHistory""" 
            df=pd.read_sql(findLastRowQuery,con = sqliteConnection)
            global RowID
            RowID=len(df)-index.row()
            df.to_sql('ImportHistory', con=sqliteConnection, if_exists='replace',index=False)
            query=f"""SELECT *
                    FROM ImportHistory
                    WHERE rowid=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(RowID,))
            sqliteConnection.commit()
            self.DateImportText.setText(df.iat[0,0])
            self.PartNameImportText.setText(df.iat[0,1])
            self.PartCodeImportText.setText(df.iat[0,2])
            self.SupplierNameText.setText(df.iat[0,3])
            self.FromText.setText(df.iat[0,4])
            self.UseForEqImportText.setText(df.iat[0,5])
            self.ModelEqImportText.setText(df.iat[0,6])
            self.QualImportText.setText(str(df.iat[0,7]))
            self.PartDesText.setText(df.iat[0,8])
        except sqlite3.Error as error:
                print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def ShowAllInforExport(self,index):
        self.DeleteExport.setEnabled(True)
        self.EditExport.setEnabled(True)
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            findLastRowQuery=f"""SELECT * FROM ExportHistory""" 
            df=pd.read_sql(findLastRowQuery,con = sqliteConnection)
            global RowIDEx
            RowIDEx=len(df)-index.row()
            df.to_sql('ExportHistory', con=sqliteConnection, if_exists='replace',index=False)
            query=f"""SELECT *
                    FROM ExportHistory
                    WHERE rowid=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(RowIDEx,))
            sqliteConnection.commit()
            self.DateExportText.setText(df.iat[0,0])
            self.PartNameExportText.setText(df.iat[0,1])
            print(self.PartNameExportText.text())
            self.PartCodeExportText.setText(df.iat[0,2])
            self.SenderText.setText(df.iat[0,3])
            self.ReceiverText.setText(df.iat[0,4])
            self.UseForEqExportText.setText(df.iat[0,5])
            self.ModelEqExportText.setText(df.iat[0,6])
            self.QualExportText.setText(str(df.iat[0,7]))
            self.WhyExportText.setText(df.iat[0,8])
        except sqlite3.Error as error:
                print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def SelectTab(self):
        if self.tabWidget.currentIndex()==2:
            self.Filter()
    
    def Filter(self):
        dfInfor=pd.DataFrame({'部品名':['Filter','Filter','Filter','Filter'],
                    '部品コード':['WGW.02.02.152','WGW.01.12.251','WGW.01.12.250','WGW.01.12.150'],
                    '部品パラメータ':['10µm 20’’(L=500×Φ60)','5µm 20’’(L=500×Φ60)','1µm 20’’(L=500×Φ60)', '1µm 10’’(L=250×Φ60'],
                    '一回交換量':[1,3,8,8],
                    '一ヶ月使用量':['',3,8,''],
                    '六ヶ月使用量':[1,'','',8]})

        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        columns = [i[1] for i in cursor.execute('PRAGMA table_info(Inventory)')]
        print(columns)
        if '部品コード' not in columns:
            cursor.close()
            sqliteConnection.close()
            return
        try:
            query=f"""SELECT 部品コード,数量
                FROM Inventory
                WHERE 部品名='Filter'"""  
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            cursor.close()
            if df.size==0:
                print("Không có filter") 
                return
            else:
                dfInfor = pd.concat([dfInfor, df], axis=1, join="inner")
                column_numbers = [x for x in range(dfInfor.shape[1])]  
                column_numbers.remove(6) 
                dfInfor=dfInfor.iloc[:, column_numbers] 
                self.model = PandasModelForFilter(dfInfor)
                self.tableView_3.setModel(self.model)

                OutOfStockDate=[]
                YesOrNo=[]
                POReleaseDate=[]
                for i in range(0,len(dfInfor)):
                    if dfInfor['一ヶ月使用量'].iat[i]=='':
                        temp1=(datetime.now()+timedelta(days=180*dfInfor['数量'].iat[i]/dfInfor['六ヶ月使用量'].iat[i])).strftime("%Y/%m/%d")
                        OutOfStockDate.append(temp1)
                        temp2=(datetime.now()+timedelta(days=180*dfInfor['数量'].iat[i]/dfInfor['六ヶ月使用量'].iat[i]-30)).strftime("%Y/%m/%d")
                        POReleaseDate.append(temp2)
                    else:
                        temp1=(datetime.now()+timedelta(days=30*dfInfor['数量'].iat[i]/dfInfor['一ヶ月使用量'].iat[i])).strftime("%Y/%m/%d")
                        OutOfStockDate.append(temp1)
                        temp2=(datetime.now()+timedelta(days=30*dfInfor['数量'].iat[i]/dfInfor['一ヶ月使用量'].iat[i]-30)).strftime("%Y/%m/%d")
                        POReleaseDate.append(temp2)

                    if dfInfor['数量'].iat[i]>=2*dfInfor['一回交換量'].iat[i]:
                        YesOrNo.append('No')
                    else:
                        YesOrNo.append('Yes')
                dfOut= pd.DataFrame(columns=['在庫切れの推定日','注文必要諾否','PO発行日'])
                dfOut['在庫切れの推定日']=OutOfStockDate
                dfOut['注文必要諾否']=YesOrNo
                dfOut['PO発行日']=POReleaseDate
                dfLeft=dfInfor.drop(['一回交換量','一ヶ月使用量','六ヶ月使用量'],axis=1,inplace=False)
                dfTable2=pd.concat([dfLeft,dfOut],axis=1)
                self.model = PandasModelForFilter(dfTable2)
                self.tableView_4.setModel(self.model)
                dfFilter=pd.concat([dfInfor,dfOut],axis=1)
                dfFilter.to_sql('Filter', con=sqliteConnection, if_exists='replace',index=False)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def SearchParts(self):
        self.SaveExcelSearchButton.setEnabled(False)
        self.SaveExcelSearchButton.hide()
        SearchText=str(self.SearchText.text())
        if SearchText=="":
            self.UpdateInventory()
            PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            return

        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT *
                FROM Inventory
                WHERE 部品名 LIKE '%'||?||'%'""" 
            df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
            sqliteConnection.commit()
            cursor.close()
            if df.size==0:
                SearchedFlat=False
                print("Không có kết quả tìm kiếm theo tên linh kiện") 
            else:
                SearchedFlat=True
                df.to_sql('SearchResult', con=sqliteConnection, if_exists='replace',index=False)
                self.model = PandasModelForImportHistoryAndSearch(df)
                self.tableViewInventory.setModel(self.model)
                self.SaveExcelSearchButton.show()
                self.SaveExcelSearchButton.setEnabled(True)
                print("Có kết quả tìm kiếm theo tên linh kiện")
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                self.SearchText.setText("")
                PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
        if SearchedFlat==False:
            try:
                query=f"""SELECT *
                        FROM Inventory
                        WHERE 部品コード LIKE '%'||?||'%'""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:
                    SearchedFlat=False
                    print("Không có kết quả tìm kiếm theo mã linh kiện")
                else:
                    SearchedFlat=True
                    df.to_sql('SearchResult', con=sqliteConnection, if_exists='replace',index=False)
                    self.model = PandasModelForImportHistoryAndSearch(df)
                    self.tableViewInventory.setModel(self.model)
                    self.SaveExcelSearchButton.show()
                    self.SaveExcelSearchButton.setEnabled(True)
                    print("Có kết quả tìm kiếm theo mã linh kiện")
            except sqlite3.Error as error:
                print("Lỗi Tìm kiếm", error)
            finally:
                if sqliteConnection:
                    self.SearchText.setText("")
                    sqliteConnection.close()
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
        if SearchedFlat==False:
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "検索結果はありません！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)

    def ResetImportFields(self):
        self.DeleteImport.setEnabled(False)
        self.EditImport.setEnabled(False)
        self.DateImportText.setText("")
        self.PartNameImportText.setText("")
        self.PartCodeImportText.setText("")
        self.SupplierNameText.setText("")
        self.FromText.setText("")
        self.UseForEqImportText.setText("")
        self.ModelEqImportText.setText("")
        self.QualImportText.setText("")
        self.PartDesText.setText("")
    
    def ResetExportFields(self):
        self.DeleteExport.setEnabled(False)
        self.EditExport.setEnabled(False)
        self.DateExportText.setText("")
        self.PartNameExportText.setText("")
        self.PartCodeExportText.setText("")
        #self.SenderText.setText("")
        self.ReceiverText.setText("")
        self.UseForEqExportText.setText("")
        self.ModelEqExportText.setText("")
        self.QualExportText.setText("")
        self.WhyExportText.setText("")

    def SaveExcelImportHistory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfHistoryImport = pd.read_sql("SELECT * FROM ImportHistory;", con=sqliteConnection)
            if not dfHistoryImport.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Inventory", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/ImportHistory_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfHistoryImport.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:I'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    #respond=self.MessageBoxOKCancel("確認","Excelファイルが正常に保存されました！\nExcelファイルを開きますか？")
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def SaveExcelExportHistory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfHistoryExport = pd.read_sql("SELECT * FROM ExportHistory;", con=sqliteConnection)
            print(dfHistoryExport)
            if not dfHistoryExport.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Inventory", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/ImportHistory_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfHistoryExport.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:I'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    #respond=self.MessageBoxOKCancel("確認","Excelファイルが正常に保存されました！\nExcelファイルを開きますか？")
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def SaveExcelInventory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfInventory = pd.read_sql("SELECT * FROM Inventory;", con=sqliteConnection)
            print(dfInventory)
            if not dfInventory.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Inventory", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/AllPartsInventory_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfInventory.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:C'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    #respond=self.MessageBoxOKCancel("確認","Excelファイルが正常に保存されました！\nExcelファイルを開きますか？")
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()
    
    def SaveExcelSearchResult(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfSearchResult = pd.read_sql("SELECT * FROM SearchResult;", con=sqliteConnection)
            if not dfSearchResult.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Inventory", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/SearchResult_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfSearchResult.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:C'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    #respond=self.MessageBoxOKCancel("確認","Excelファイルが正常に保存されました！\nExcelファイルを開きますか？")
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def SaveExcelFilter(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfSearchResult = pd.read_sql("SELECT * FROM Filter;", con=sqliteConnection)
            if not dfSearchResult.size==0:
                defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
                des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Inventory", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
                if not des_folderpath=="":
                    FileName="/Filter_Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                    full_path=des_folderpath+FileName
                    dfSearchResult.to_excel(full_path, index = False)
                    #định dạng lại file excel
                    wb = load_workbook(full_path)
                    ws = wb.active
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    lastRow=ws.max_row
                    for row in ws['A1:J'+str(lastRow)]:
                        for cell in row:
                            cell.border = border
                    for idx, col in enumerate(ws.columns, 1):
                        ws.column_dimensions[get_column_letter(idx)].auto_size = True
                    wb.save(full_path)
                    #hỏi người dùng có muốn xem file excel
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                    #respond=self.MessageBoxOKCancel("確認","Excelファイルが正常に保存されました！\nExcelファイルを開きますか？")
                    if respond==QMessageBox.Ok:
                        os.chdir(des_folderpath)
                        cut_string=FileName.split('/')
                        new_string = cut_string[1]
                        cmd="start excel.exe "+new_string
                        #os.system(cmd)
                        subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def AdExportHistory(self):
        if self.SenderText.text()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        DateInput=self.DateExportText.text()
        #kiểm tra part name nhập có đúng
        PartName=self.PartNameExportText.text()
        if PartName=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品名を入力してください！"))
            return
        else:
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT 部品名 FROM Inventory""" 
                cursor.execute(query)
                sqliteConnection.commit()
                final_result = [i[0] for i in cursor.fetchall()]
                cursor.close()
                if not PartName in final_result:
                    self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "データベースにこの部品名が存在しません！"))
                    self.PartNameExportText.setText("")
                    PyQt5.QtCore.QTimer.singleShot(0, self.PartNameExportText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()  
        #kiểm tra part code nhập có đúng
        PartCode=self.PartCodeExportText.text()
        if PartCode=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品コードを入力してください！"))
            return
        else:
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT 部品コード FROM Inventory""" 
                cursor.execute(query)
                sqliteConnection.commit()
                final_result = [i[0] for i in cursor.fetchall()]
                cursor.close()
                if not PartCode in final_result:
                    self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "データベースにこの部品コードが存在しません！"))
                    self.PartCodeExportText.setText("")
                    PyQt5.QtCore.QTimer.singleShot(0, self.PartCodeExportText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()  

        Sender=self.SenderText.text()
        if Sender=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "取出人を入力してください！"))
            return
        Receiver=self.ReceiverText.text()
        if Receiver=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "受取人を入力してください！"))
            return
        EqUse=self.UseForEqExportText.text()
        EqID=self.ModelEqExportText.text()
        Qual=self.QualExportText.text()
        if Qual=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "数量を入力してください！"))
            return
        else:
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT 数量 FROM Inventory WHERE 部品コード=?""" 
                data_tuples=(PartCode,)
                cursor.execute(query,data_tuples)
                sqliteConnection.commit()
                inventoryQuantity=cursor.fetchone()[0]
                cursor.close()
                if inventoryQuantity<int(Qual):
                    self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "在庫量は輸出されるのに十分ではありません!"))
                    self.QualExportText.setText("")
                    PyQt5.QtCore.QTimer.singleShot(0, self.QualExportText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 

        Reason=self.WhyExportText.text()
        if Reason=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "理由を入力してください！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            #cập nhật vô QTableView Export History
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_insert_new_record_query= f"""INSERT INTO ExportHistory VALUES (?,?,?,?,?,?,?,?,?)""" 
                data_tuple = (DateInput,PartName,PartCode,Sender,Receiver,EqUse,EqID,Qual,Reason)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "部品輸出成功！"))
            except sqlite3.Error as error:
                print("Lỗi ", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    self.UpdateExportHistory()
            #tính tồn kho cập nhật vô QTableView Search
            self.UpdateInventory()
            #cập nhật logger thông tin linh kiện xuất
            try:
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                newOperate='This user exported from the warehouse the part named '+PartName+',code '+PartCode+',quatity '+Qual+' pcs.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 
            #reset các trường đã nhập
        self.ResetExportFields()

    def UpdateExportHistory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT * FROM ExportHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            cursor.close()

        except sqlite3.Error as error:
            print("Lỗi ", error)
        finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    if not df.size==0:
                        reversed_df = df.iloc[::-1]
                        self.model = PandasModelForExportHistory(reversed_df)
                        self.tableViewExport.setModel(self.model)
                    else:
                        emptydf = pd.DataFrame(columns=['日付','部品名','部品コード','取出人','受取人','使用設備','設備機種','数量','理由'])
                        self.model = PandasModelForExportHistory(emptydf)
                        self.tableViewExport.setModel(self.model)

    def AdImportHistory(self):
        if self.SenderText.text()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "このアカウントはデータを編集する許可がありません！"))
            return
        DateInput=self.DateImportText.text()

        PartName=self.PartNameImportText.text()
        if PartName=="":
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品名を入力してください！"))
            return

        PartCode=self.PartCodeImportText.text()
        if PartCode=="":
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "部品コードを入力してください！"))
            return
 
        Supplier=self.SupplierNameText.text()
        From=self.FromText.text()
        EqUse=self.UseForEqImportText.text()
        EqID=self.ModelEqImportText.text()
        Qual=self.QualImportText.text()
        if Qual=="":
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "エラー"),QtWidgets.QApplication.translate("Inventory", "数量を入力してください！"))
            return
        Des=self.PartDesText.text()
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            #cập nhật vô QTableView Import History
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_insert_new_record_query= f"""INSERT INTO ImportHistory VALUES (?,?,?,?,?,?,?,?,?)""" 
                data_tuple = (DateInput,PartName,PartCode,Supplier,From,EqUse,EqID,Qual,Des)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Inventory", "確認"),QtWidgets.QApplication.translate("Inventory", "部品輸入成功！"))
            except sqlite3.Error as error:
                print("Lỗi ", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    self.UpdateImportHistory()
            #tính tồn kho cập nhật vô QTableView Search
            self.UpdateInventory()
            #cập nhật logger thông tin linh kiện nhập
            try:
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                newOperate='This user imported to the warehouse the part named '+PartName+',code '+PartCode+',quatity '+Qual+' pcs.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 
            #cập nhật tên linh kiện mới và mã linh kiện mới vào database
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor = sqliteConnection.cursor()
            try:
                query=f"""SELECT partName FROM PartList""" 
                cursor.execute(query)
                final_result = [i[0] for i in cursor.fetchall()]
                if not PartName in final_result:
                    query=f"""INSERT INTO PartList VALUES (?)"""
                    cursor.execute(query,(PartName,))
                
                query=f"""SELECT partCode FROM PartCodeList""" 
                cursor.execute(query)
                final_result = [i[0] for i in cursor.fetchall()]
                if not PartName in final_result:
                    query=f"""INSERT INTO PartCodeList VALUES (?)"""
                    cursor.execute(query,(PartCode,))
                sqliteConnection.commit()
                cursor.close()	
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()    
        #reset các trường đã nhập
        self.ResetImportFields()

    def UpdateInventory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfImport = pd.read_sql("SELECT 部品名,部品コード,SUM(数量) FROM ImportHistory GROUP BY 部品コード ORDER BY 部品名;", con=sqliteConnection)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        try:
            dfExport = pd.read_sql("SELECT 部品コード,SUM(数量) FROM ExportHistory GROUP BY 部品コード ORDER BY 部品名;", con=sqliteConnection)
            dfInventory=dfImport.merge(dfExport,how='outer',on='部品コード')
            dfInventory=dfInventory.fillna(0)
            dfInventory['数量']=dfInventory['SUM(数量)_x']-dfInventory['SUM(数量)_y']
            dfInventory.drop(dfInventory.columns[[2,3]],axis=1,inplace=True)
            dfInventory.数量 = dfInventory.数量.astype(int) 
            dfInventory.to_sql('Inventory', con=sqliteConnection, if_exists='replace',index=False)
            self.modelInventory = PandasModelForImportHistoryAndSearch(dfInventory)
            self.tableViewInventory.setModel(self.modelInventory)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

    def UpdateImportHistory(self):
        sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT * FROM ImportHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            cursor.close()
            if not df.size==0:
                reversed_df = df.iloc[::-1]
                self.model = PandasModelForImportHistoryAndSearch(reversed_df)
                self.tableViewImport.setModel(self.model)
            else:
                        emptydf = pd.DataFrame(columns=['日付','部品名','部品コード','供給者','原産地','使用設備','設備機種','数量','部品パラメータ'])
                        self.model = PandasModelForExportHistory(emptydf)
                        self.tableViewImport.setModel(self.model)

        except sqlite3.Error as error:
            print("Lỗi ", error)
        finally:
                if sqliteConnection:
                    sqliteConnection.close()

class Login(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(Login,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/Login.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        
        self.InitialValue()
        self.Signal_Slot()

    def __del__(self):
        print("hàm hủy của Login đã được gọi!")

    def InitialValue(self):
        self.trans = PyQt5.QtCore.QTranslator(self)
        options = [
            ("日本語", ""),
            ("English", "jpn-eng"),
        ]
        for i, (text, lang) in enumerate(options):
            self.comboBox.addItem(text)
            self.comboBox.setItemData(i, lang)
        self.retranslateUi()

        PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)
        #self.UserNameText.setPlaceholderText("ユーザネーム") 
        #self.PassText.setPlaceholderText("パスワード") 
        self.PassText.setEchoMode(QtWidgets.QLineEdit.Password)
        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # 
        self.showTime()

    @PyQt5.QtCore.pyqtSlot(int)
    def ChangeLanguage(self, index):
        data = self.comboBox.itemData(index)
        print(data)
        if data=="jpn-eng":
            self.trans.load('jpn-eng.qm')
            QtWidgets.QApplication.instance().installTranslator(self.trans)
            print("change language to english!")

        elif data=="":
            if self.trans is not None:
                QtWidgets.QApplication.instance().removeTranslator(self.trans)
                print("remove language")

    def changeEvent(self, event):
        if event.type() == PyQt5.QtCore.QEvent.LanguageChange:
            print("run changeEvent")
            self.retranslateUi()
        super(Login, self).changeEvent(event)
    
    def retranslateUi(self):
        print("run retranslateUi")
        self.label.setText(QtWidgets.QApplication.translate("Login", "貴方の母国語を選択して"))
        self.label_4.setText(QtWidgets.QApplication.translate("Login", "設備技術課の管理システムへようこそ！"))
        self.LoginButton.setText(QtWidgets.QApplication.translate("Login", "登錄"))
        self.ResetLogin.setText(QtWidgets.QApplication.translate("Login", "再設定"))
        self.ExitButton.setText(QtWidgets.QApplication.translate("Login", "終了"))
        self.UserNameText.setPlaceholderText(QtWidgets.QApplication.translate("Login", "ユーザー名を入力")) 
        self.PassText.setPlaceholderText(QtWidgets.QApplication.translate("Login", "パスワードを入力"))
    
    def Signal_Slot(self):
        self.LoginButton.clicked.connect(self.GoToDashBoard)
        self.ExitButton.clicked.connect(self.ExitSystem)
        self.ResetLogin.clicked.connect(self.ResetLoginInfor)
        self.PassText.returnPressed.connect(self.GoToDashBoard)
        self.shortcut_exit = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_exit.activated.connect(self.ExitSystemQuick)
        self.comboBox.currentIndexChanged.connect(self.ChangeLanguage)
        self.shortcut_selectEnglish = QShortcut(QKeySequence('Ctrl+E'), self)
        self.shortcut_selectEnglish.activated.connect(self.SelectEnglish)
        self.shortcut_selectJapanese = QShortcut(QKeySequence('Ctrl+J'), self)
        self.shortcut_selectJapanese.activated.connect(self.SelectJapanese)

    def SelectEnglish(self):
        self.comboBox.setCurrentText("English")
        PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)
    
    def SelectJapanese(self):
        self.comboBox.setCurrentText("日本語")
        PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)

    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_clock.setText(displayTxt)
 
    def ResetLoginInfor(self):
        self.UserNameText.setText("")
        self.PassText.setText("")
        PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)

    def GoToDashBoard(self):
        if self.UserNameText.text()=="" and self.PassText.text()=="":
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Login", "エラー"),QtWidgets.QApplication.translate("Login", "ログイン情報を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)
        elif self.UserNameText.text()=="admin" and self.PassText.text()=="admin" and os.getlogin()=="LIN JIN YU" or \
            self.UserNameText.text()=="guest" and self.PassText.text()=="guest" or \
            self.UserNameText.text()=="gkm577" and self.PassText.text()=="smv577" or \
            self.UserNameText.text()=="gkm327" and self.PassText.text()=="smv327" or \
            self.UserNameText.text()=="gkm124" and self.PassText.text()=="smv124" or \
            self.UserNameText.text()=="gkm068" and self.PassText.text()=="smv068" or \
            self.UserNameText.text()=="gkm690" and self.PassText.text()=="smv690" or \
            self.UserNameText.text()=="gkm648" and self.PassText.text()=="smv648" or \
            self.UserNameText.text()=="gkm673" and self.PassText.text()=="smv673":

            self.UserName=self.UserNameText.text()
            print(self.UserName)
            if self.UserName=="guest" or self.UserName=="admin":
                pass
            else:
                self.UserName=os.getlogin()
            print(self.UserName)
            self.UserNameText.setText("")
            self.PassText.setText("")
            PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)

            DB=DashBoard()
            widget.addWidget(DB)
            widget.setFixedHeight(604)
            widget.setFixedWidth(800)
            widget.setWindowTitle("Dashboard Window")
            widget.setWindowIcon(QIcon(":icon/DashBoard.png"))
            widget.setCurrentWidget(DB)
            widget.move(screen_center(widget))
            
            #cập nhật file log
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            try:
                sqlite_insert_logger_query= f"""INSERT INTO UserHistory VALUES (?,?,?,?,?)""" 
                data_tuple = (self.UserName,socket.gethostbyname(socket.gethostname()),datetime.now().strftime("%d/%m/%Y %H:%M:%S"),'','')
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi ", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()

        else:
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Login", "エラー"),QtWidgets.QApplication.translate("Login", "ログイン情報が間違っています！"))
            self.UserNameText.setText("")
            self.PassText.setText("")
            PyQt5.QtCore.QTimer.singleShot(0, self.UserNameText.setFocus)

class Equipment(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(Equipment,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/EqControl.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
        ####################Khởi tạo mặc định############################
        self.InitialValue()
        ###################các sự kiện tương tác####################
        self.Signal_Slot()

    def __del__(self):
        print("hàm hủy của Equipment đã được gọi!")

    def Signal_Slot(self):
        self.radioButton_1p.toggled.connect(self.Value1P)
        self.radioButton_3p.toggled.connect(self.Value3P)
        self.checkBox_Air.stateChanged.connect(self.state_changed_Air)
        self.checkBox_N2.stateChanged.connect(self.state_changed_N2)
        self.checkBox_VAC.stateChanged.connect(self.state_changed_VAC)
        self.checkBox_O2.stateChanged.connect(self.state_changed_O2)
        self.checkBox_Arg.stateChanged.connect(self.state_changed_Arg)
        self.checkBox_Oil.stateChanged.connect(self.state_changed_Oil)
        self.checkBox_Emision.stateChanged.connect(self.state_changed_Emision)
        self.checkBox_CleanWater.stateChanged.connect(self.state_changed_CleanWater)
        self.checkBox_CoolingWater.stateChanged.connect(self.state_changed_CoolingWater)
        self.SearchButton.clicked.connect(self.SearchEquipment)
        self.SearchText.returnPressed.connect(self.SearchEquipment)
        self.NewFormButton.clicked.connect(self.AdNewForm)
        self.DeleteButton.clicked.connect(self.DeleteForm)
        self.EditButton.clicked.connect(self.EditForm)
        self.InputButton.clicked.connect(self.AdNewHistory)
        self.ResetButton_2.clicked.connect(self.ResetFieldMaintenanceHistory)
        self.ResetButton.clicked.connect(self.ResetFieldForm)
        self.Button_Image.clicked.connect(self.UploadImage)
        self.Button_ImageFactory.clicked.connect(self.UploadFactoryImage)
        self.tableView_Search.pressed.connect(self.DisplayAllInfor)
        self.SaveExcelButton.clicked.connect(self.SaveExcel)
        self.ReturnButton.clicked.connect(self.BackToDashBoard)
        self.ExitButton.clicked.connect(self.ExitSystem)
        self.shortcut_exit = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_exit.activated.connect(self.ExitSystemQuick)
        self.shortcut_returnToDashBoard = QShortcut(QKeySequence('Ctrl+B'), self)
        self.shortcut_returnToDashBoard.activated.connect(self.BackToDashBoard)

    def InitialValue(self):
        #DỊCH NÚT NHẤN
        self.label_2.setText(QtWidgets.QApplication.translate("Equipment", "CM/SMT設備台帳表"))

        self.tabWidget.setTabText(0,QtWidgets.QApplication.translate("Equipment", "一般情報(1)")) 
        self.tabWidget.setTabText(1,QtWidgets.QApplication.translate("Equipment", "一般情報(2)"))
        self.tabWidget.setTabText(2,QtWidgets.QApplication.translate("Equipment", "写真"))
        self.tabWidget.setTabText(3,QtWidgets.QApplication.translate("Equipment", "定期検査")) 

        self.NewFormButton.setText(QtWidgets.QApplication.translate("Equipment", "新規作成"))
        self.DeleteButton.setText(QtWidgets.QApplication.translate("Equipment", "消去"))
        self.EditButton.setText(QtWidgets.QApplication.translate("Equipment", "編集"))
        self.ExitButton.setText(QtWidgets.QApplication.translate("Equipment", "終了"))
        self.ResetButton.setText(QtWidgets.QApplication.translate("Equipment", "再設定"))
        self.ReturnButton.setText(QtWidgets.QApplication.translate("Equipment", "戻る"))
        self.SearchButton.setText(QtWidgets.QApplication.translate("Equipment", "検索"))
        self.Button_Image.setText(QtWidgets.QApplication.translate("Equipment", "写真アップロード"))
        self.Button_ImageFactory.setText(QtWidgets.QApplication.translate("Equipment", "写真アップロード"))
        self.InputButton.setText(QtWidgets.QApplication.translate("Equipment", "確認"))
        self.ResetButton_2.setText(QtWidgets.QApplication.translate("Equipment", "再設定"))
        #DỊCH THÔNG TIN TAB 1
        self.EqName.setText(QtWidgets.QApplication.translate("Equipment", "設備名"))
        self.Model.setText(QtWidgets.QApplication.translate("Equipment", "機種"))
        self.serial.setText(QtWidgets.QApplication.translate("Equipment", "製造番号"))
        self.date.setText(QtWidgets.QApplication.translate("Equipment", "製造日"))
        self.money.setText(QtWidgets.QApplication.translate("Equipment", "物価"))
        self.maker.setText(QtWidgets.QApplication.translate("Equipment", "メーカ"))
        self.length.setText(QtWidgets.QApplication.translate("Equipment", "長(mm)"))
        self.width.setText(QtWidgets.QApplication.translate("Equipment", "幅(mm)"))
        self.height.setText(QtWidgets.QApplication.translate("Equipment", "高(mm)"))
        self.weight.setText(QtWidgets.QApplication.translate("Equipment", "重(kg)"))
        self.DateInput.setText(QtWidgets.QApplication.translate("Equipment", "導入日"))
        self.invoice.setText(QtWidgets.QApplication.translate("Equipment", "インボイス"))
        self.support.setText(QtWidgets.QApplication.translate("Equipment", "サポート"))
        self.EqID.setText(QtWidgets.QApplication.translate("Equipment", "設備ID"))
        self.proType.setText(QtWidgets.QApplication.translate("Equipment", "財産種"))
        self.FixProNumSharp.setText(QtWidgets.QApplication.translate("Equipment", "固定財産番号(Sharp)"))
        self.process.setText(QtWidgets.QApplication.translate("Equipment", "工程"))
        self.location.setText(QtWidgets.QApplication.translate("Equipment", "位置"))
        self.FixProNumSMV.setText(QtWidgets.QApplication.translate("Equipment", "固定財産番号(SMV)"))
        self.vol.setText(QtWidgets.QApplication.translate("Equipment", "電圧(V)"))
        self.current.setText(QtWidgets.QApplication.translate("Equipment", "電流(A)"))
        self.power.setText(QtWidgets.QApplication.translate("Equipment", "パワー数"))
        self.ratedPower.setText(QtWidgets.QApplication.translate("Equipment", "規制(KVA)"))
        self.consumedPower.setText(QtWidgets.QApplication.translate("Equipment", "消費(KVA)"))
        #DỊCH THÔNG TIN TAB 2
        self.label_qual.setText(QtWidgets.QApplication.translate("Equipment", "数量"))
        self.label_pressure.setText(QtWidgets.QApplication.translate("Equipment", "圧力"))
        self.label_flow.setText(QtWidgets.QApplication.translate("Equipment", "流量"))
        self.checkBox_Emision.setText(QtWidgets.QApplication.translate("Equipment", "排気"))
        self.checkBox_CleanWater.setText(QtWidgets.QApplication.translate("Equipment", "純水"))
        self.checkBox_CoolingWater.setText(QtWidgets.QApplication.translate("Equipment", "冷却水"))
        #DỊCH THÔNG TIN TAB 3
        self.label_3.setText(QtWidgets.QApplication.translate("Equipment", "設備の実際写真"))
        self.label_4.setText(QtWidgets.QApplication.translate("Equipment", "現場に設備の実際写真"))
        self.label_image.setText(QtWidgets.QApplication.translate("Equipment", "一枚写真を選択してください。"))
        self.label_FactoryImage.setText(QtWidgets.QApplication.translate("Equipment", "一枚写真を選択してください。"))
        #DỊCH THÔNG TIN TAB 4
        self.label_5.setText(QtWidgets.QApplication.translate("Equipment", "定期点検・修理・改善"))
        self.label_6.setText(QtWidgets.QApplication.translate("Equipment", "日付"))
        self.label_7.setText(QtWidgets.QApplication.translate("Equipment", "箇所"))
        self.label_8.setText(QtWidgets.QApplication.translate("Equipment", "実施内容"))
        self.label_9.setText(QtWidgets.QApplication.translate("Equipment", "担当者"))
        self.label_10.setText(QtWidgets.QApplication.translate("Equipment", "確認者"))

        self.DeleteButton.setEnabled(False)
        self.EditButton.setEnabled(False)
        self.SaveExcelButton.setEnabled(False)
        
        self.lineEdit_1p.setEnabled(False)
        self.lineEdit_3p.setEnabled(False)
        self.lineEdit_qualAir.setEnabled(False)
        self.lineEdit_phiAir.setEnabled(False)
        self.lineEdit_pressureAir.setEnabled(False)
        self.lineEdit_flowAir.setEnabled(False)
        self.lineEdit_qualN2.setEnabled(False)
        self.lineEdit_phiN2.setEnabled(False)
        self.lineEdit_pressureN2.setEnabled(False)
        self.lineEdit_flowN2.setEnabled(False)
        self.lineEdit_qualVAC.setEnabled(False)
        self.lineEdit_phiVAC.setEnabled(False)
        self.lineEdit_pressureVAC.setEnabled(False)
        self.lineEdit_flowVAC.setEnabled(False)
        self.lineEdit_qualO2.setEnabled(False)
        self.lineEdit_phiO2.setEnabled(False)
        self.lineEdit_pressureO2.setEnabled(False)
        self.lineEdit_flowO2.setEnabled(False)
        self.lineEdit_qualArg.setEnabled(False)
        self.lineEdit_phiArg.setEnabled(False)
        self.lineEdit_pressureArg.setEnabled(False)
        self.lineEdit_flowArg.setEnabled(False)
        self.lineEdit_qualOil.setEnabled(False)
        self.lineEdit_phiOil.setEnabled(False)
        self.lineEdit_pressureOil.setEnabled(False)
        self.lineEdit_flowOil.setEnabled(False)
        self.lineEdit_qualEmision.setEnabled(False)
        self.lineEdit_phiEmision.setEnabled(False)
        self.lineEdit_pressureEmision.setEnabled(False)
        self.lineEdit_flowEmision.setEnabled(False)
        self.lineEdit_qualCleanWater.setEnabled(False)
        self.lineEdit_phiCleanWater.setEnabled(False)
        self.lineEdit_pressureCleanWater.setEnabled(False)
        self.lineEdit_flowCleanWater.setEnabled(False)
        self.lineEdit_qualCoolingWater.setEnabled(False)
        self.lineEdit_phiCoolingWater.setEnabled(False)
        self.lineEdit_pressureCoolingWater.setEnabled(False)
        self.lineEdit_flowCoolingWater.setEnabled(False)
        validatorMoney = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_money.setValidator(validatorMoney)
        validatorLength = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_length.setValidator(validatorLength)
        validatorWidth = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_width.setValidator(validatorWidth)
        validatorHeight = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_height.setValidator(validatorHeight)
        validatorWeight = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_weight.setValidator(validatorWeight)
        validator1p = QRegExpValidator(QRegExp("[0-9]{15}"))
        self.lineEdit_1p.setValidator(validator1p)
        validator3p = QRegExpValidator(QRegExp("[0-9]{15}"))
        self.lineEdit_3p.setValidator(validator3p)
        validatorCurrent = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_current.setValidator(validatorCurrent)
        validatorRatedPower = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_ratedPower.setValidator(validatorRatedPower)
        validatorConsumedPower = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_consumedPower.setValidator(validatorConsumedPower)
        validatorQualAir = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualAir.setValidator(validatorQualAir)
        validatorPhiAir = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiAir.setValidator(validatorPhiAir)
        validatorPressureAir = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureAir.setValidator(validatorPressureAir)
        validatorFlowAir = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowAir.setValidator(validatorFlowAir)
        validatorQualN2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualN2.setValidator(validatorQualN2)
        validatorPhiN2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiN2.setValidator(validatorPhiN2)
        validatorPressureN2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureN2.setValidator(validatorPressureN2)
        validatorFlowN2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowN2.setValidator(validatorFlowN2)
        validatorQualVAC = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualVAC.setValidator(validatorQualVAC)
        validatorPhiVAC = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiVAC.setValidator(validatorPhiVAC)
        validatorPressureVAC = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureVAC.setValidator(validatorPressureVAC)
        validatorFlowVAC = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowVAC.setValidator(validatorFlowVAC)
        validatorQualO2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualO2.setValidator(validatorQualO2)
        validatorPhiO2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiO2.setValidator(validatorPhiO2)
        validatorPressureO2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureO2.setValidator(validatorPressureO2)
        validatorFlowO2 = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowO2.setValidator(validatorFlowO2)
        validatorQualArg = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualArg.setValidator(validatorQualArg)
        validatorPhiArg = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiArg.setValidator(validatorPhiArg)
        validatorPressureArg = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureArg.setValidator(validatorPressureArg)
        validatorFlowArg = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowArg.setValidator(validatorFlowArg)
        validatorQualOil = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualOil.setValidator(validatorQualOil)
        validatorPhiOil = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiOil.setValidator(validatorPhiOil)
        validatorPressureOil = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureOil.setValidator(validatorPressureOil)
        validatorFlowOil = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowOil.setValidator(validatorFlowOil)
        validatorQualEmision = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualEmision.setValidator(validatorQualEmision)
        validatorPhiEmision = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiEmision.setValidator(validatorPhiEmision)
        validatorPressureEmision = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureEmision.setValidator(validatorPressureEmision)
        validatorFlowEmision = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowEmision.setValidator(validatorFlowEmision)
        validatorQualCleanWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualCleanWater.setValidator(validatorQualCleanWater)
        validatorPhiCleanWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiCleanWater.setValidator(validatorPhiCleanWater)
        validatorPressureCleanWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureCleanWater.setValidator(validatorPressureCleanWater)
        validatorFlowCleanWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowCleanWater.setValidator(validatorFlowCleanWater)
        validatorQualCoolingWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_qualCoolingWater.setValidator(validatorQualCoolingWater)
        validatorPhiCoolingWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_phiCoolingWater.setValidator(validatorPhiCoolingWater)
        validatorPressureCoolingWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_pressureCoolingWater.setValidator(validatorPressureCoolingWater)
        validatorFlowCoolingWater = QRegExpValidator(QRegExp("[0-9-.]{15}"))
        self.lineEdit_flowCoolingWater.setValidator(validatorFlowCoolingWater)
        dataEquipmentName=['Solder Printer','Chip Mounter','PKG Mounter','Reflow','Auto Visual Inspection machine',\
            'Flux Cleaner','Loader','Unloader','PKG Removal jig','Conveyor']
        self.lineEdit_EqName.setCompleter(QCompleter(dataEquipmentName))
        dataProcessName=['SMT','CM']
        self.lineEdit_process.setCompleter(QCompleter(dataProcessName))
        #self.lineEdit_date_2.setDate(datetime.now().date())
        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # update sau mỗi giây
        self.showTime()

    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_clock.setText(displayTxt)

    def EditForm(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "このアカウントはデータを編集する許可がありません！"))
            return

        EquipmentName=self.lineEdit_EqName.text()
        if EquipmentName=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "設備名を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_EqName.setFocus)
            return
        Model=self.lineEdit_Model.text()
        Serial=self.lineEdit_serial.text()
        if Serial=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "製造番号を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_serial.setFocus)
            return
        DateOfManufacture=self.lineEdit_date.text()
        Money=self.lineEdit_money.text()
        Maker=self.lineEdit_maker.text()
        Length=self.lineEdit_length.text()
        Width=self.lineEdit_width.text()
        Height=self.lineEdit_height.text()
        Weight=self.lineEdit_weight.text()
        DateInput=self.lineEdit_DateInput.text()
        Invoice=self.lineEdit_invoice.text()
        Supporter=self.lineEdit_supporter.text()
        EquipmentID=self.lineEdit_EqID.text()
        PropertyType=self.lineEdit_proType.text()
        FixedPropertyNumberSharp=self.lineEdit_FixProNumSharp.text()
        ProcessName=self.lineEdit_process.text()
        Location=self.lineEdit_location.text()
        FixedPropertyNumberSMV=self.lineEdit_FixProNumSMV.text()
        if self.radioButton_1p.isChecked():
            Voltage=self.lineEdit_1p.text()
            VolType=1
        elif self.radioButton_3p.isChecked():
            Voltage=self.lineEdit_3p.text()
            VolType=3
        else:
            VolType=""
            Voltage=""
        Current=self.lineEdit_current.text()
        RatedPower=self.lineEdit_ratedPower.text()
        ConsumedPower=self.lineEdit_consumedPower.text()
        ###kiểm tra air
        if self.checkBox_Air.isChecked():
            AirCheck=1
        else:
            AirCheck=0
        qualAir=self.lineEdit_qualAir.text()
        phiAir=self.lineEdit_phiAir.text()
        pressureAir=self.lineEdit_pressureAir.text()
        flowAir=self.lineEdit_flowAir.text()
        if AirCheck==1 and qualAir=="" and phiAir=="" and pressureAir=="" and flowAir=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Air CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra N2
        if self.checkBox_N2.isChecked():
            N2Check=1
        else:
            N2Check=0
        qualN2=self.lineEdit_qualN2.text()
        phiN2=self.lineEdit_phiN2.text()
        pressureN2=self.lineEdit_pressureN2.text()
        flowN2=self.lineEdit_flowN2.text()
        if N2Check==1 and qualN2=="" and phiN2=="" and pressureN2=="" and flowN2=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "N2 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra VAC
        if self.checkBox_VAC.isChecked():
            VACCheck=1
        else:
            VACCheck=0
        qualVAC=self.lineEdit_qualVAC.text()
        phiVAC=self.lineEdit_phiVAC.text()
        pressureVAC=self.lineEdit_pressureVAC.text()
        flowVAC=self.lineEdit_flowVAC.text()
        if VACCheck==1 and qualVAC=="" and phiVAC=="" and pressureVAC=="" and flowVAC=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "VAC CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra O2
        if self.checkBox_O2.isChecked():
            O2Check=1
        else:
            O2Check=0
        qualO2=self.lineEdit_qualO2.text()
        phiO2=self.lineEdit_phiO2.text()
        pressureO2=self.lineEdit_pressureO2.text()
        flowO2=self.lineEdit_flowO2.text()
        if O2Check==1 and qualO2=="" and phiO2=="" and pressureO2=="" and flowO2=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "O2 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra Arg
        if self.checkBox_Arg.isChecked():
            ArgCheck=1
        else:
            ArgCheck=0
        qualArg=self.lineEdit_qualArg.text()
        phiArg=self.lineEdit_phiArg.text()
        pressureArg=self.lineEdit_pressureArg.text()
        flowArg=self.lineEdit_flowArg.text()
        if ArgCheck==1 and qualArg=="" and phiArg=="" and pressureArg=="" and flowArg=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Arg CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra Oil
        if self.checkBox_Oil.isChecked():
            OilCheck=1
        else:
            OilCheck=0
        qualOil=self.lineEdit_qualOil.text()
        phiOil=self.lineEdit_phiOil.text()
        pressureOil=self.lineEdit_pressureOil.text()
        flowOil=self.lineEdit_flowOil.text()
        if OilCheck==1 and qualOil=="" and phiOil=="" and pressureOil=="" and flowOil=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Oil CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra khí thải
        if self.checkBox_Emision.isChecked():
            EmisionCheck=1
        else:
            EmisionCheck=0
        qualEmision=self.lineEdit_qualEmision.text()
        phiEmision=self.lineEdit_phiEmision.text()
        pressureEmision=self.lineEdit_pressureEmision.text()
        flowEmision=self.lineEdit_flowEmision.text()
        if EmisionCheck==1 and qualEmision=="" and phiEmision=="" and pressureEmision=="" and flowEmision=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "排気 CheckBoxを選択しましたが何も入力していません！"))
            return
        #kiểm tra nước sạch
        if self.checkBox_CleanWater.isChecked():
            CleanWaterCheck=1
        else:
            CleanWaterCheck=0
        qualCleanWater=self.lineEdit_qualCleanWater.text()
        phiCleanWater=self.lineEdit_phiCleanWater.text()
        pressureCleanWater=self.lineEdit_pressureCleanWater.text()
        flowCleanWater=self.lineEdit_flowCleanWater.text()
        if CleanWaterCheck==1 and qualCleanWater=="" and phiCleanWater=="" and pressureCleanWater=="" and flowCleanWater=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "純水 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra nước làm mát
        if self.checkBox_CoolingWater.isChecked():
            CoolingWaterCheck=1
        else:
            CoolingWaterCheck=0
        qualCoolingWater=self.lineEdit_qualCoolingWater.text()
        phiCoolingWater=self.lineEdit_phiCoolingWater.text()
        pressureCoolingWater=self.lineEdit_pressureCoolingWater.text()
        flowCoolingWater=self.lineEdit_flowCoolingWater.text()
        if CoolingWaterCheck==1 and qualCoolingWater=="" and phiCoolingWater=="" and pressureCoolingWater=="" and flowCoolingWater=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "冷却水 CheckBoxを選択しましたが何も入力していません！"))
            return

        HasEditedPhoto=False
        try:
            imagePath
            Photo = self.convertToBinaryData(imagePath)
            HasEditedPhoto=True
        except:
            HasEditedPhoto=False
        HasEditedFactoryPhoto=False
        try:
            FactoryImagePath
            InFactoryPhoto = self.convertToBinaryData(FactoryImagePath)
            HasEditedFactoryPhoto=True
        except:
            HasEditedFactoryPhoto=False
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "本気ですか？"))
        if respond==QMessageBox.Ok:	
            try:
                sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
                cursor = sqliteConnection.cursor()
                if HasEditedPhoto==True and HasEditedFactoryPhoto==True:
                    sqlite_update_record_query= f"""UPDATE Equipment 
                                    SET 設備名=?,機種=?,製造番号=?,製造日=?,物価=?,メーカー=?,長=?,幅=?,高=?,重=?,導入日=?,インボイス=?,サポート=?,設備ID=?,財産種=?,固定財産番号Sharp=?,工程=?,位置=?,
                                    固定財産番号SMV=?,電圧種=?,電圧値=?,電流=?,規制パワー=?,消費パワー=?,AirCheck=?,数量Air=?,TupeAir=?,圧力Air=?,流量Air=?,N2Check=?,数量N2=?,TupeN2=?,
                                    圧力N2=?,流量N2=?,VACCheck=?,数量VAC=?,TupeVAC=?,圧力VAC=?,流量VAC=?,O2Check=?,数量O2=?,TupeO2=?,圧力O2=?,流量O2=?,ArgCheck=?,数量Arg=?,TupeArg=?,
                                    圧力Arg=?,流量Arg=?,OilCheck=?,数量Oil=?,TupeOil=?,圧力Oil=?,流量Oil=?,EmisionCheck=?,数量Emision=?,TupeEmision=?,圧力Emision=?,流量Emision=?,
                                    CleanWaterCheck=?,数量CleanWater=?,TupeCleanWater=?,圧力CleanWater=?,流量CleanWater=?,CoolingWaterCheck=?,数量CoolingWater=?,TupeCoolingWater=?,圧力CoolingWater=?,
                                    流量CoolingWater=?,Image=?,ImageFactory=?
                                    WHERE RowID=?"""
                    data_tuple = (EquipmentName, Model, Serial, DateOfManufacture, Money, Maker, Length, Width, Height, Weight, DateInput, \
                                                Invoice, Supporter, EquipmentID, PropertyType, FixedPropertyNumberSharp, ProcessName, Location, FixedPropertyNumberSMV, VolType, \
                                                Voltage, Current, RatedPower, ConsumedPower, AirCheck, qualAir, phiAir, pressureAir, flowAir, N2Check, qualN2, phiN2, pressureN2, \
                                                flowN2, VACCheck, qualVAC, phiVAC, pressureVAC, flowVAC, O2Check, qualO2, phiO2, pressureO2, flowO2, ArgCheck, qualArg, phiArg, \
                                                pressureArg, flowArg, OilCheck, qualOil, phiOil, pressureOil, flowOil, EmisionCheck, qualEmision, phiEmision, pressureEmision, \
                                                flowEmision, CleanWaterCheck, qualCleanWater, phiCleanWater, pressureCleanWater, flowCleanWater, CoolingWaterCheck, qualCoolingWater, \
                                                phiCoolingWater, pressureCoolingWater, flowCoolingWater,Photo,InFactoryPhoto, key)
                elif HasEditedPhoto==True:
                    sqlite_update_record_query= f"""UPDATE Equipment 
                                    SET 設備名=?,機種=?,製造番号=?, 製造日=?,物価=?,メーカー=?,長=?,幅=?,高=?,重=?,導入日=?,インボイス=?,サポート=?,設備ID=?,財産種=?,固定財産番号Sharp=?,工程=?,位置=?,
                                    固定財産番号SMV=?,電圧種=?,電圧値=?,電流=?,規制パワー=?,消費パワー=?,AirCheck=?,数量Air=?,TupeAir=?,圧力Air=?,流量Air=?,N2Check=?,数量N2=?,TupeN2=?,
                                    圧力N2=?,流量N2=?,VACCheck=?,数量VAC=?,TupeVAC=?,圧力VAC=?,流量VAC=?,O2Check=?,数量O2=?,TupeO2=?,圧力O2=?,流量O2=?,ArgCheck=?,数量Arg=?,TupeArg=?,
                                    圧力Arg=?,流量Arg=?,OilCheck=?,数量Oil=?,TupeOil=?,圧力Oil=?,流量Oil=?,EmisionCheck=?,数量Emision=?,TupeEmision=?,圧力Emision=?,流量Emision=?,
                                    CleanWaterCheck=?,数量CleanWater=?,TupeCleanWater=?,圧力CleanWater=?,流量CleanWater=?,CoolingWaterCheck=?,数量CoolingWater=?,TupeCoolingWater=?,圧力CoolingWater=?,
                                    流量CoolingWater=?,Image=?
                                    WHERE RowID=?"""
                    data_tuple = (EquipmentName, Model, Serial, DateOfManufacture, Money, Maker, Length, Width, Height, Weight, DateInput, \
                                                Invoice, Supporter, EquipmentID, PropertyType, FixedPropertyNumberSharp, ProcessName, Location, FixedPropertyNumberSMV, VolType, \
                                                Voltage, Current, RatedPower, ConsumedPower, AirCheck, qualAir, phiAir, pressureAir, flowAir, N2Check, qualN2, phiN2, pressureN2, \
                                                flowN2, VACCheck, qualVAC, phiVAC, pressureVAC, flowVAC, O2Check, qualO2, phiO2, pressureO2, flowO2, ArgCheck, qualArg, phiArg, \
                                                pressureArg, flowArg, OilCheck, qualOil, phiOil, pressureOil, flowOil, EmisionCheck, qualEmision, phiEmision, pressureEmision, \
                                                flowEmision, CleanWaterCheck, qualCleanWater, phiCleanWater, pressureCleanWater, flowCleanWater, CoolingWaterCheck, qualCoolingWater, \
                                                phiCoolingWater, pressureCoolingWater, flowCoolingWater,Photo, key)
                elif HasEditedFactoryPhoto==True:
                    sqlite_update_record_query= f"""UPDATE Equipment 
                                    SET 設備名=?,機種=?,製造番号=?, 製造日=?,物価=?,メーカー=?,長=?,幅=?,高=?,重=?,導入日=?,インボイス=?,サポート=?,設備ID=?,財産種=?,固定財産番号Sharp=?,工程=?,位置=?,
                                    固定財産番号SMV=?,電圧種=?,電圧値=?,電流=?,規制パワー=?,消費パワー=?,AirCheck=?,数量Air=?,TupeAir=?,圧力Air=?,流量Air=?,N2Check=?,数量N2=?,TupeN2=?,
                                    圧力N2=?,流量N2=?,VACCheck=?,数量VAC=?,TupeVAC=?,圧力VAC=?,流量VAC=?,O2Check=?,数量O2=?,TupeO2=?,圧力O2=?,流量O2=?,ArgCheck=?,数量Arg=?,TupeArg=?,
                                    圧力Arg=?,流量Arg=?,OilCheck=?,数量Oil=?,TupeOil=?,圧力Oil=?,流量Oil=?,EmisionCheck=?,数量Emision=?,TupeEmision=?,圧力Emision=?,流量Emision=?,
                                    CleanWaterCheck=?,数量CleanWater=?,TupeCleanWater=?,圧力CleanWater=?,流量CleanWater=?,CoolingWaterCheck=?,数量CoolingWater=?,TupeCoolingWater=?,圧力CoolingWater=?,
                                    流量CoolingWater=?,ImageFactory=?
                                    WHERE RowID=?"""
                    data_tuple = (EquipmentName, Model, Serial, DateOfManufacture, Money, Maker, Length, Width, Height, Weight, DateInput, \
                                                Invoice, Supporter, EquipmentID, PropertyType, FixedPropertyNumberSharp, ProcessName, Location, FixedPropertyNumberSMV, VolType, \
                                                Voltage, Current, RatedPower, ConsumedPower, AirCheck, qualAir, phiAir, pressureAir, flowAir, N2Check, qualN2, phiN2, pressureN2, \
                                                flowN2, VACCheck, qualVAC, phiVAC, pressureVAC, flowVAC, O2Check, qualO2, phiO2, pressureO2, flowO2, ArgCheck, qualArg, phiArg, \
                                                pressureArg, flowArg, OilCheck, qualOil, phiOil, pressureOil, flowOil, EmisionCheck, qualEmision, phiEmision, pressureEmision, \
                                                flowEmision, CleanWaterCheck, qualCleanWater, phiCleanWater, pressureCleanWater, flowCleanWater, CoolingWaterCheck, qualCoolingWater, \
                                                phiCoolingWater, pressureCoolingWater, flowCoolingWater,InFactoryPhoto, key)
                else:
                    sqlite_update_record_query= f"""UPDATE Equipment 
                                    SET 設備名=?,機種=?,製造番号=?, 製造日=?,物価=?,メーカー=?,長=?,幅=?,高=?,重=?,導入日=?,インボイス=?,サポート=?,設備ID=?,財産種=?,固定財産番号Sharp=?,工程=?,位置=?,
                                    固定財産番号SMV=?,電圧種=?,電圧値=?,電流=?,規制パワー=?,消費パワー=?,AirCheck=?,数量Air=?,TupeAir=?,圧力Air=?,流量Air=?,N2Check=?,数量N2=?,TupeN2=?,
                                    圧力N2=?,流量N2=?,VACCheck=?,数量VAC=?,TupeVAC=?,圧力VAC=?,流量VAC=?,O2Check=?,数量O2=?,TupeO2=?,圧力O2=?,流量O2=?,ArgCheck=?,数量Arg=?,TupeArg=?,
                                    圧力Arg=?,流量Arg=?,OilCheck=?,数量Oil=?,TupeOil=?,圧力Oil=?,流量Oil=?,EmisionCheck=?,数量Emision=?,TupeEmision=?,圧力Emision=?,流量Emision=?,
                                    CleanWaterCheck=?,数量CleanWater=?,TupeCleanWater=?,圧力CleanWater=?,流量CleanWater=?,CoolingWaterCheck=?,数量CoolingWater=?,TupeCoolingWater=?,圧力CoolingWater=?,
                                    流量CoolingWater=?
                                    WHERE RowID=?"""
                    data_tuple = (EquipmentName, Model, Serial, DateOfManufacture, Money, Maker, Length, Width, Height, Weight, DateInput, \
                                                Invoice, Supporter, EquipmentID, PropertyType, FixedPropertyNumberSharp, ProcessName, Location, FixedPropertyNumberSMV, VolType, \
                                                Voltage, Current, RatedPower, ConsumedPower, AirCheck, qualAir, phiAir, pressureAir, flowAir, N2Check, qualN2, phiN2, pressureN2, \
                                                flowN2, VACCheck, qualVAC, phiVAC, pressureVAC, flowVAC, O2Check, qualO2, phiO2, pressureO2, flowO2, ArgCheck, qualArg, phiArg, \
                                                pressureArg, flowArg, OilCheck, qualOil, phiOil, pressureOil, flowOil, EmisionCheck, qualEmision, phiEmision, pressureEmision, \
                                                flowEmision, CleanWaterCheck, qualCleanWater, phiCleanWater, pressureCleanWater, flowCleanWater, CoolingWaterCheck, qualCoolingWater, \
                                                phiCoolingWater, pressureCoolingWater, flowCoolingWater, key)
                cursor.execute(sqlite_update_record_query, data_tuple)
                sqliteConnection.commit()
                print("Cập nhật database thành công")
                cursor.close()
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "データベース更新成功！"))
                PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                self.UpdateSearchUI()
                self.ResetFieldForm()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")

    def DisplayAllInfor(self,index):
        self.ResetFieldForm()
        self.ResetFieldMaintenanceHistory
        self.DeleteButton.setEnabled(True)
        self.EditButton.setEnabled(True)
        self.SaveExcelButton.setEnabled(True)
        global key
        key=index.sibling(index.row(),0).data()
        global serial
        serial=index.sibling(index.row(),2).data()
        sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
        sortRowIDQuery=f"""SELECT * FROM Equipment"""
        df=df=pd.read_sql(sortRowIDQuery,con = sqliteConnection)
        df.to_sql('Equipment', con=sqliteConnection, if_exists='replace',index=False)

        try:
            query=f"""SELECT *
                        FROM Equipment
                        WHERE RowID=?""" 
            df=pd.read_sql(query,con = sqliteConnection,params=(key,))
            sqliteConnection.commit()
        except sqlite3.Error as error:
                print("Lỗi", error)
        finally:
            if sqliteConnection:
                self.lineEdit_EqName.setText(df.iat[0,0])
                self.lineEdit_Model.setText(df.iat[0,1])
                self.lineEdit_serial.setText(df.iat[0,2])
                self.lineEdit_date.setText(df.iat[0,3])
                self.lineEdit_money.setText(df.iat[0,4])
                self.lineEdit_maker.setText(df.iat[0,5])
                self.lineEdit_length.setText(str(df.iat[0,6]))
                self.lineEdit_width.setText(str(df.iat[0,7]))
                self.lineEdit_height.setText(str(df.iat[0,8]))
                self.lineEdit_weight.setText(str(df.iat[0,9]))
                self.lineEdit_DateInput.setText(df.iat[0,10])
                self.lineEdit_invoice.setText(df.iat[0,11])
                self.lineEdit_supporter.setText(df.iat[0,12])
                self.lineEdit_EqID.setText(df.iat[0,13])
                self.lineEdit_proType.setText(df.iat[0,14])
                self.lineEdit_FixProNumSharp.setText(df.iat[0,15])
                self.lineEdit_process.setText(df.iat[0,16])
                self.lineEdit_location.setText(df.iat[0,17])
                self.lineEdit_FixProNumSMV.setText(df.iat[0,18])
                if not df.iat[0,19]=="":
                    if int(df.iat[0,19])==3:
                        self.radioButton_3p.setChecked(True)
                        self.lineEdit_3p.setText(str(df.iat[0,20]))
                    elif int(df.iat[0,19])==1:
                        self.radioButton_1p.setChecked(True)
                        self.lineEdit_1p.setText(str(df.iat[0,20]))
                else:
                    self.radioButton_1p.setChecked(False)
                    self.radioButton_3p.setChecked(False)
                    self.lineEdit_1p.setText("")
                    self.lineEdit_3p.setText("")

                self.lineEdit_current.setText(str(df.iat[0,21]))
                self.lineEdit_ratedPower.setText(str(df.iat[0,22]))
                self.lineEdit_consumedPower.setText(str(df.iat[0,23]))
                if df.iat[0,24]==1:
                    self.checkBox_Air.setChecked(True)
                    self.lineEdit_qualAir.setText(str(df.iat[0,25]))
                    self.lineEdit_phiAir.setText(str(df.iat[0,26]))
                    self.lineEdit_pressureAir.setText(str(df.iat[0,27]))
                    self.lineEdit_flowAir.setText(str(df.iat[0,28]))
                if df.iat[0,29]==1:
                    self.checkBox_N2.setChecked(True)
                    self.lineEdit_qualN2.setText(str(df.iat[0,30]))
                    self.lineEdit_phiN2.setText(str(df.iat[0,31]))
                    self.lineEdit_pressureN2.setText(str(df.iat[0,32]))
                    self.lineEdit_flowN2.setText(str(df.iat[0,33]))
                if df.iat[0,34]==1:
                    self.checkBox_VAC.setChecked(True)
                    self.lineEdit_qualVAC.setText(str(df.iat[0,35]))
                    self.lineEdit_phiVAC.setText(str(df.iat[0,36]))
                    self.lineEdit_pressureVAC.setText(str(df.iat[0,37]))
                    self.lineEdit_flowVAC.setText(str(df.iat[0,38]))
                if df.iat[0,39]==1:
                    self.checkBox_O2.setChecked(True)
                    self.lineEdit_qualO2.setText(str(df.iat[0,40]))
                    self.lineEdit_phiO2.setText(str(df.iat[0,41]))
                    self.lineEdit_pressureO2.setText(str(df.iat[0,42]))
                    self.lineEdit_flowO2.setText(str(df.iat[0,43]))
                if df.iat[0,44]==1:
                    self.checkBox_Arg.setChecked(True)
                    self.lineEdit_qualArg.setText(str(df.iat[0,45]))
                    self.lineEdit_phiArg.setText(str(df.iat[0,46]))
                    self.lineEdit_pressureArg.setText(str(df.iat[0,47]))
                    self.lineEdit_flowArg.setText(str(df.iat[0,48]))
                if df.iat[0,49]==1:
                    self.checkBox_Oil.setChecked(True)
                    self.lineEdit_qualOil.setText(str(df.iat[0,50]))
                    self.lineEdit_phiOil.setText(str(df.iat[0,51]))
                    self.lineEdit_pressureOil.setText(str(df.iat[0,52]))
                    self.lineEdit_flowOil.setText(str(df.iat[0,53]))
                if df.iat[0,54]==1:
                    self.checkBox_Emision.setChecked(True)
                    self.lineEdit_qualEmision.setText(str(df.iat[0,55]))
                    self.lineEdit_phiEmision.setText(str(df.iat[0,56]))
                    self.lineEdit_pressureEmision.setText(str(df.iat[0,57]))
                    self.lineEdit_flowEmision.setText(str(df.iat[0,58]))
                if df.iat[0,59]==1:
                    self.checkBox_CleanWater.setChecked(True)
                    self.lineEdit_qualCleanWater.setText(str(df.iat[0,60]))
                    self.lineEdit_phiCleanWater.setText(str(df.iat[0,61]))
                    self.lineEdit_pressureCleanWater.setText(str(df.iat[0,62]))
                    self.lineEdit_flowCleanWater.setText(str(df.iat[0,63]))
                if df.iat[0,64]==1:
                    self.checkBox_CoolingWater.setChecked(True)
                    self.lineEdit_qualCoolingWater.setText(str(df.iat[0,65]))
                    self.lineEdit_phiCoolingWater.setText(str(df.iat[0,66]))
                    self.lineEdit_pressureCoolingWater.setText(str(df.iat[0,67]))
                    self.lineEdit_flowCoolingWater.setText(str(df.iat[0,68]))

                try:
                    cursor = sqliteConnection.cursor()
                    sqlite_earch_image_query= f"""SELECT Image FROM Equipment WHERE RowID=?"""
                    data=(key,)
                    cursor.execute(sqlite_earch_image_query,data)
                    sqliteConnection.commit()
                    result=cursor.fetchone()
                    image=result[0]
                    pixmap=PyQt5.QtGui.QPixmap()
                    pixmap.loadFromData(image)
                    self.label_image.setPixmap(pixmap)
                    cursor.close()
                except sqlite3.Error as error:
                    print("Lỗi", error)

                try:
                    cursor = sqliteConnection.cursor()
                    sqlite_earch_image_query= f"""SELECT ImageFactory FROM Equipment WHERE RowID=?"""
                    data=(key,)
                    cursor.execute(sqlite_earch_image_query,data)
                    sqliteConnection.commit()
                    result=cursor.fetchone()
                    image=result[0]
                    pixmap=PyQt5.QtGui.QPixmap()
                    pixmap.loadFromData(image)
                    self.label_FactoryImage.setPixmap(pixmap)
                    cursor.close()
                except sqlite3.Error as error:
                    print("Lỗi", error)

            try:
                query=f"""SELECT * FROM MaintenanceHistory WHERE 製造番号=?"""
                df=pd.read_sql(query,con=sqliteConnection,params=(serial,))
                sqliteConnection.commit()
                cursor.close()
                if not df.size==0:
                    reversed_df = df.iloc[::-1]
                    self.model = MaintenanceHistoryTableModel(reversed_df)
                    self.tableView_Check.setModel(self.model)
                else:
                    df = pd.DataFrame(columns=['設備名','製造番号','日付','箇所','実施内容','担当者','確認者'])
                    self.model = MaintenanceHistoryTableModel(df)
                    self.tableView_Check.setModel(self.model)
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                sqliteConnection.close()
            return

    def Value1P(self):
        self.lineEdit_1p.setEnabled(True)
        self.lineEdit_3p.setEnabled(False)
        self.lineEdit_3p.setText("")

    def Value3P(self):
        self.lineEdit_1p.setEnabled(False)
        self.lineEdit_3p.setEnabled(True)
        self.lineEdit_1p.setText("")
    
    def state_changed_Air(self):
        if self.checkBox_Air.isChecked():
            self.lineEdit_qualAir.setEnabled(True)
            self.lineEdit_phiAir.setEnabled(True)
            self.lineEdit_pressureAir.setEnabled(True)
            self.lineEdit_flowAir.setEnabled(True)
        else:
            self.lineEdit_qualAir.setEnabled(False)
            self.lineEdit_phiAir.setEnabled(False)
            self.lineEdit_pressureAir.setEnabled(False)
            self.lineEdit_flowAir.setEnabled(False)
            self.lineEdit_qualAir.setText("")
            self.lineEdit_phiAir.setText("")
            self.lineEdit_pressureAir.setText("")
            self.lineEdit_flowAir.setText("")

    def state_changed_N2(self):
        if self.checkBox_N2.isChecked():
            self.lineEdit_qualN2.setEnabled(True)
            self.lineEdit_phiN2.setEnabled(True)
            self.lineEdit_pressureN2.setEnabled(True)
            self.lineEdit_flowN2.setEnabled(True)
        else:
            self.lineEdit_qualN2.setEnabled(False)
            self.lineEdit_phiN2.setEnabled(False)
            self.lineEdit_pressureN2.setEnabled(False)
            self.lineEdit_flowN2.setEnabled(False)
            self.lineEdit_qualN2.setText("")
            self.lineEdit_phiN2.setText("")
            self.lineEdit_pressureN2.setText("")
            self.lineEdit_flowN2.setText("")

    def state_changed_VAC(self):
        if self.checkBox_VAC.isChecked():
            self.lineEdit_qualVAC.setEnabled(True)
            self.lineEdit_phiVAC.setEnabled(True)
            self.lineEdit_pressureVAC.setEnabled(True)
            self.lineEdit_flowVAC.setEnabled(True)
        else:
            self.lineEdit_qualVAC.setEnabled(False)
            self.lineEdit_phiVAC.setEnabled(False)
            self.lineEdit_pressureVAC.setEnabled(False)
            self.lineEdit_flowVAC.setEnabled(False)
            self.lineEdit_qualVAC.setText("")
            self.lineEdit_phiVAC.setText("")
            self.lineEdit_pressureVAC.setText("")
            self.lineEdit_flowVAC.setText("")

    def state_changed_O2(self):
        if self.checkBox_O2.isChecked():
            self.lineEdit_qualO2.setEnabled(True)
            self.lineEdit_phiO2.setEnabled(True)
            self.lineEdit_pressureO2.setEnabled(True)
            self.lineEdit_flowO2.setEnabled(True)
        else:
            self.lineEdit_qualO2.setEnabled(False)
            self.lineEdit_phiO2.setEnabled(False)
            self.lineEdit_pressureO2.setEnabled(False)
            self.lineEdit_flowO2.setEnabled(False)
            self.lineEdit_qualO2.setText("")
            self.lineEdit_phiO2.setText("")
            self.lineEdit_pressureO2.setText("")
            self.lineEdit_flowO2.setText("")

    def state_changed_Arg(self):
        if self.checkBox_Arg.isChecked():
            self.lineEdit_qualArg.setEnabled(True)
            self.lineEdit_phiArg.setEnabled(True)
            self.lineEdit_pressureArg.setEnabled(True)
            self.lineEdit_flowArg.setEnabled(True)
        else:
            self.lineEdit_qualArg.setEnabled(False)
            self.lineEdit_phiArg.setEnabled(False)
            self.lineEdit_pressureArg.setEnabled(False)
            self.lineEdit_flowArg.setEnabled(False)
            self.lineEdit_qualArg.setText("")
            self.lineEdit_phiArg.setText("")
            self.lineEdit_pressureArg.setText("")
            self.lineEdit_flowArg.setText("")
    
    def state_changed_Oil(self):
        if self.checkBox_Oil.isChecked():
            self.lineEdit_qualOil.setEnabled(True)
            self.lineEdit_phiOil.setEnabled(True)
            self.lineEdit_pressureOil.setEnabled(True)
            self.lineEdit_flowOil.setEnabled(True)
        else:
            self.lineEdit_qualOil.setEnabled(False)
            self.lineEdit_phiOil.setEnabled(False)
            self.lineEdit_pressureOil.setEnabled(False)
            self.lineEdit_flowOil.setEnabled(False)
            self.lineEdit_qualOil.setText("")
            self.lineEdit_phiOil.setText("")
            self.lineEdit_pressureOil.setText("")
            self.lineEdit_flowOil.setText("")

    def state_changed_Emision(self):
        if self.checkBox_Emision.isChecked():
            self.lineEdit_qualEmision.setEnabled(True)
            self.lineEdit_phiEmision.setEnabled(True)
            self.lineEdit_pressureEmision.setEnabled(True)
            self.lineEdit_flowEmision.setEnabled(True)
        else:
            self.lineEdit_qualEmision.setEnabled(False)
            self.lineEdit_phiEmision.setEnabled(False)
            self.lineEdit_pressureEmision.setEnabled(False)
            self.lineEdit_flowEmision.setEnabled(False)
            self.lineEdit_qualEmision.setText("")
            self.lineEdit_phiEmision.setText("")
            self.lineEdit_pressureEmision.setText("")
            self.lineEdit_flowEmision.setText("")

    def state_changed_CleanWater(self):
        if self.checkBox_CleanWater.isChecked():
            self.lineEdit_qualCleanWater.setEnabled(True)
            self.lineEdit_phiCleanWater.setEnabled(True)
            self.lineEdit_pressureCleanWater.setEnabled(True)
            self.lineEdit_flowCleanWater.setEnabled(True)
        else:
            self.lineEdit_qualCleanWater.setEnabled(False)
            self.lineEdit_phiCleanWater.setEnabled(False)
            self.lineEdit_pressureCleanWater.setEnabled(False)
            self.lineEdit_flowCleanWater.setEnabled(False)
            self.lineEdit_qualCleanWater.setText("")
            self.lineEdit_phiCleanWater.setText("")
            self.lineEdit_pressureCleanWater.setText("")
            self.lineEdit_flowCleanWater.setText("")

    def state_changed_CoolingWater(self):
        if self.checkBox_CoolingWater.isChecked():
            self.lineEdit_qualCoolingWater.setEnabled(True)
            self.lineEdit_phiCoolingWater.setEnabled(True)
            self.lineEdit_pressureCoolingWater.setEnabled(True)
            self.lineEdit_flowCoolingWater.setEnabled(True)
        else:
            self.lineEdit_qualCoolingWater.setEnabled(False)
            self.lineEdit_phiCoolingWater.setEnabled(False)
            self.lineEdit_pressureCoolingWater.setEnabled(False)
            self.lineEdit_flowCoolingWater.setEnabled(False)
            self.lineEdit_qualCoolingWater.setText("")
            self.lineEdit_phiCoolingWater.setText("")
            self.lineEdit_pressureCoolingWater.setText("")
            self.lineEdit_flowCoolingWater.setText("")

    def UploadImage(self):
        self.label_image.setPixmap(PyQt5.QtGui.QPixmap())
        defaultfolder='C:\\Users\\{0}\\Downloads'.format(os.getlogin())
        fname = QtWidgets.QFileDialog.getOpenFileName(self, QtWidgets.QApplication.translate("Equipment", "一枚写真を選択してください!"), defaultfolder, "Image files (*.jpg *.png)")
        global imagePath
        imagePath = fname[0]
        if imagePath=="":
            imagePath=DefaultPhotoPath
        pixmap = PyQt5.QtGui.QPixmap(imagePath)
        self.label_image.setPixmap(PyQt5.QtGui.QPixmap(pixmap))

    def UploadFactoryImage(self):
        self.label_FactoryImage.setPixmap(PyQt5.QtGui.QPixmap())
        defaultfolder='C:\\Users\\{0}\\Downloads'.format(os.getlogin())
        fname = QtWidgets.QFileDialog.getOpenFileName(self, QtWidgets.QApplication.translate("Equipment", "一枚写真を選択してください!"), defaultfolder, "Image files (*.jpg *.png)")
        global FactoryImagePath
        FactoryImagePath = fname[0]
        if FactoryImagePath=="":
            FactoryImagePath=DefaultPhotoPath
        pixmap = PyQt5.QtGui.QPixmap(FactoryImagePath)
        self.label_FactoryImage.setPixmap(PyQt5.QtGui.QPixmap(pixmap))

    def convertToBinaryData(self,filename):
        with open(filename, 'rb') as file:
            blobData = file.read()
        return blobData

    def writeTofile(self,data, filename):
        with open(filename, 'wb') as file:
            file.write(data)
        print("Stored blob data into: ", filename, "\n")

    def AdNewHistory(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "このアカウントはデータを編集する許可がありません！"))
            return
        #Date=str(self.lineEdit_date_2.date().toPyDate())
        Date=self.lineEdit_date_2.text()
        Location=self.lineEdit_location_2.text()
        Content=self.lineEdit_content.text()
        PIC=self.lineEdit_pic.text()
        Confirmer=self.lineEdit_confirmer.text()
        if not Date=="" and not Location=="" and not Content=="" and not PIC=="" and not Confirmer=="":
            EquipmentName=self.lineEdit_EqName.text()
            Serial=self.lineEdit_serial.text()
            try:
                sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
                cursor = sqliteConnection.cursor()
                sqlite_insert_new_record_query= f"""INSERT INTO MaintenanceHistory VALUES (?,?,?,?,?,?,?)"""
                data_tuple = (EquipmentName, Serial, Date, Location, Content, PIC, Confirmer)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                print("Bản ghi lịch sử bảo trì thiết bị đã được thêm thành công vào database")
                cursor.close()
                df = pd.read_sql("SELECT * FROM MaintenanceHistory WHERE 製造番号=?", con=sqliteConnection,params=(Serial,))
                reversed_df = df.iloc[::-1]
                self.model = MaintenanceHistoryTableModel(reversed_df)
                self.tableView_Check.setModel(self.model)
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "情報入力成功！"))
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    self.lineEdit_date_2.setText("")
                    self.lineEdit_location_2.setText("")
                    self.lineEdit_content.setText("")
                    self.lineEdit_pic.setText("")
                    self.lineEdit_confirmer.setText("")
                    sqliteConnection.close()

        else:
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "十分に情報を入力してください！"))
    
    def SearchEquipment(self):
        self.ResetFieldForm()
        self.ResetFieldMaintenanceHistory()
        SearchText=str(self.SearchText.text())
        if SearchText=="":
            self.UpdateSearchUI()
            PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            return
        print(SearchText)
        sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                FROM Equipment
                WHERE 設備名 LIKE '%'||?||'%'""" 
            df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
            sqliteConnection.commit()
            cursor.close()
            if df.size==0:
                SearchedFlat=False
                print("Không có kết quả tìm kiếm theo tên thiết bị")
            else:
                SearchedFlat=True
                self.model = MaintenanceHistoryTableModel(df)
                self.tableView_Search.setModel(self.model)
                print("Có kết quả tìm kiếm theo tên thiết bị")
                PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                return
        except sqlite3.Error as error:
            print("Lỗi Tìm kiếm", error)
        finally:
                if sqliteConnection:
                    self.SearchText.setText("")
                    
        if SearchedFlat==False:
            try:
                query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                        FROM Equipment
                        WHERE 製造番号 LIKE '%'||?||'%'""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:
                    SearchedFlat=False
                    print("Không có kết quả tìm kiếm theo số serial")
                else:
                    SearchedFlat=True
                    self.model = MaintenanceHistoryTableModel(df)
                    self.tableView_Search.setModel(self.model)
                    print("Có kết quả tìm kiếm theo số serial")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi Tìm kiếm", error)
            finally:
                if sqliteConnection:
                    self.SearchText.setText("")
                    
        if SearchedFlat==False:                  
            try:
                query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                        FROM Equipment
                        WHERE 固定財産番号Sharp LIKE '%'||?||'%'""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:
                    SearchedFlat=False
                    print("Không có kết quả tìm kiếm theo mã số tài sản cố định Sharp")
                else:
                    SearchedFlat=True
                    self.model = MaintenanceHistoryTableModel(df)
                    self.tableView_Search.setModel(self.model)
                    print("Có kết quả tìm kiếm theo mã số tài sản cố định Sharp")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi Tìm kiếm", error)
            finally:
                if sqliteConnection:
                    self.SearchText.setText("")

        if SearchedFlat==False:            
            try:
                query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                        FROM Equipment
                        WHERE 固定財産番号SMV LIKE '%'||?||'%'""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:
                    SearchedFlat=False
                    print("Không có kết quả tìm kiếm theo mã số tài sản cố định SMV")
                else:
                    SearchedFlat=True
                    self.model = MaintenanceHistoryTableModel(df)
                    self.tableView_Search.setModel(self.model)
                    print("Có kết quả tìm kiếm theo mã số tài sản cố định SMV")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi Tìm kiếm", error)
            finally:
                if sqliteConnection:
                    self.SearchText.setText("")

        if SearchedFlat==False:  
            try:
                query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                        FROM Equipment
                        WHERE 設備ID LIKE '%'||?||'%'""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(SearchText,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:
                    SearchedFlat=False
                    print("Không có kết quả tìm kiếm theo id thiết bị")
                else:
                    SearchedFlat=True
                    self.model = MaintenanceHistoryTableModel(df)
                    self.tableView_Search.setModel(self.model)
                    print("Có kết quả tìm kiếm theo id thiết bị")
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
                    return
            except sqlite3.Error as error:
                print("Lỗi Tìm kiếm", error)
            finally:
                if sqliteConnection:
                    self.SearchText.setText("")
                    sqliteConnection.close()
        
        if SearchedFlat==False:
                    self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "検索結果はありません！"))
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)

    def ResetFieldMaintenanceHistory(self):
        self.lineEdit_date_2.setText("")
        self.lineEdit_location_2.setText("")
        self.lineEdit_content.setText("")
        self.lineEdit_pic.setText("")
        self.lineEdit_confirmer.setText("")

    def ResetFieldForm(self):
        self.DeleteButton.setEnabled(False)
        self.EditButton.setEnabled(False)
        self.SaveExcelButton.setEnabled(False)

        self.lineEdit_EqName.setText("")
        self.lineEdit_Model.setText("")
        self.lineEdit_serial.setText("")
        self.lineEdit_date.setText("")
        self.lineEdit_money.setText("")
        self.lineEdit_maker.setText("")
        self.lineEdit_length.setText("")
        self.lineEdit_width.setText("")
        self.lineEdit_height.setText("")
        self.lineEdit_weight.setText("")
        self.lineEdit_DateInput.setText("")
        self.lineEdit_invoice.setText("")
        self.lineEdit_supporter.setText("")
        self.lineEdit_EqID.setText("")
        self.lineEdit_proType.setText("")
        self.lineEdit_FixProNumSharp.setText("")
        self.lineEdit_process.setText("")
        self.lineEdit_location.setText("")
        self.lineEdit_FixProNumSMV.setText("")
        self.lineEdit_1p.setText("")
        self.lineEdit_3p.setText("")
        self.radioButton_1p.setChecked(False)
        self.radioButton_3p.setChecked(False)
        self.lineEdit_current.setText("")
        self.lineEdit_ratedPower.setText("")
        self.lineEdit_consumedPower.setText("")
        self.lineEdit_qualAir.setText("")
        self.lineEdit_phiAir.setText("")
        self.lineEdit_pressureAir.setText("")
        self.lineEdit_flowAir.setText("")
        self.lineEdit_qualN2.setText("")
        self.lineEdit_phiN2.setText("")
        self.lineEdit_pressureN2.setText("")
        self.lineEdit_flowN2.setText("")
        self.lineEdit_qualVAC.setText("")
        self.lineEdit_phiVAC.setText("")
        self.lineEdit_pressureVAC.setText("")
        self.lineEdit_flowVAC.setText("")
        self.lineEdit_qualO2.setText("")
        self.lineEdit_phiO2.setText("")
        self.lineEdit_pressureO2.setText("")
        self.lineEdit_flowO2.setText("")
        self.lineEdit_qualArg.setText("")
        self.lineEdit_phiArg.setText("")
        self.lineEdit_pressureArg.setText("")
        self.lineEdit_flowArg.setText("")
        self.lineEdit_qualOil.setText("")
        self.lineEdit_phiOil.setText("")
        self.lineEdit_pressureOil.setText("")
        self.lineEdit_flowOil.setText("")
        self.lineEdit_qualEmision.setText("")
        self.lineEdit_phiEmision.setText("")
        self.lineEdit_pressureEmision.setText("")
        self.lineEdit_flowEmision.setText("")
        self.lineEdit_qualCleanWater.setText("")
        self.lineEdit_phiCleanWater.setText("")
        self.lineEdit_pressureCleanWater.setText("")
        self.lineEdit_flowCleanWater.setText("")
        self.lineEdit_qualCoolingWater.setText("")
        self.lineEdit_phiCoolingWater.setText("")
        self.lineEdit_pressureCoolingWater.setText("")
        self.lineEdit_flowCoolingWater.setText("")
        self.checkBox_Air.setChecked(False)
        self.checkBox_N2.setChecked(False)
        self.checkBox_VAC.setChecked(False)
        self.checkBox_O2.setChecked(False)
        self.checkBox_Arg.setChecked(False)
        self.checkBox_Oil.setChecked(False)
        self.checkBox_Emision.setChecked(False)
        self.checkBox_CleanWater.setChecked(False)
        self.checkBox_CoolingWater.setChecked(False)
        self.label_image.setPixmap(PyQt5.QtGui.QPixmap())
        self.label_FactoryImage.setPixmap(PyQt5.QtGui.QPixmap())
        df = pd.DataFrame(columns=['設備名','製造番号','日付','箇所','実施内容','担当者','確認者'])
        self.model = MaintenanceHistoryTableModel(df)
        self.tableView_Check.setModel(self.model)
        PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)

    def AdNewForm(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "このアカウントはデータを編集する許可がありません！"))
            return
        sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT 製造番号 FROM Equipment""" 
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            cursor.close()
            IsDuplicated=self.lineEdit_serial.text()
            if IsDuplicated in df.values:
                print("KHông thể trùng dữ liệu số serial")
                self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "製造番号が重複しています！"))
                PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 

        EquipmentName=self.lineEdit_EqName.text()
        if EquipmentName=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "設備名を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_EqName.setFocus)
            return
        Model=self.lineEdit_Model.text()
        if Model=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "機種を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_Model.setFocus)
            return
        Serial=self.lineEdit_serial.text()
        if Serial=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "製造番号を入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_serial.setFocus)
            return
        DateOfManufacture=self.lineEdit_date.text()
        Money=self.lineEdit_money.text()
        Maker=self.lineEdit_maker.text()
        Length=self.lineEdit_length.text()
        Width=self.lineEdit_width.text()
        Height=self.lineEdit_height.text()
        Weight=self.lineEdit_weight.text()
        DateInput=self.lineEdit_DateInput.text()
        Invoice=self.lineEdit_invoice.text()
        Supporter=self.lineEdit_supporter.text()
        EquipmentID=self.lineEdit_EqID.text()
        if EquipmentID=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "設備IDを入力してください！"))
            PyQt5.QtCore.QTimer.singleShot(0, self.lineEdit_EqID.setFocus)
            return
        PropertyType=self.lineEdit_proType.text()
        FixedPropertyNumberSharp=self.lineEdit_FixProNumSharp.text()
        ProcessName=self.lineEdit_process.text()
        Location=self.lineEdit_location.text()
        FixedPropertyNumberSMV=self.lineEdit_FixProNumSMV.text()
        if self.radioButton_1p.isChecked():
            Voltage=self.lineEdit_1p.text()
            VolType=1
        elif self.radioButton_3p.isChecked():
            Voltage=self.lineEdit_3p.text()
            VolType=3
        else:
            VolType=""
            Voltage=""
        Current=self.lineEdit_current.text()
        RatedPower=self.lineEdit_ratedPower.text()
        ConsumedPower=self.lineEdit_consumedPower.text()
        ###kiểm tra air
        if self.checkBox_Air.isChecked():
            AirCheck=1
        else:
            AirCheck=0
        qualAir=self.lineEdit_qualAir.text()
        phiAir=self.lineEdit_phiAir.text()
        pressureAir=self.lineEdit_pressureAir.text()
        flowAir=self.lineEdit_flowAir.text()
        if AirCheck==1 and qualAir=="" and phiAir=="" and pressureAir=="" and flowAir=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Air CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra N2
        if self.checkBox_N2.isChecked():
            N2Check=1
        else:
            N2Check=0
        qualN2=self.lineEdit_qualN2.text()
        phiN2=self.lineEdit_phiN2.text()
        pressureN2=self.lineEdit_pressureN2.text()
        flowN2=self.lineEdit_flowN2.text()
        if N2Check==1 and qualN2=="" and phiN2=="" and pressureN2=="" and flowN2=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "N2 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra VAC
        if self.checkBox_VAC.isChecked():
            VACCheck=1
        else:
            VACCheck=0
        qualVAC=self.lineEdit_qualVAC.text()
        phiVAC=self.lineEdit_phiVAC.text()
        pressureVAC=self.lineEdit_pressureVAC.text()
        flowVAC=self.lineEdit_flowVAC.text()
        if VACCheck==1 and qualVAC=="" and phiVAC=="" and pressureVAC=="" and flowVAC=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "VAC CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra O2
        if self.checkBox_O2.isChecked():
            O2Check=1
        else:
            O2Check=0
        qualO2=self.lineEdit_qualO2.text()
        phiO2=self.lineEdit_phiO2.text()
        pressureO2=self.lineEdit_pressureO2.text()
        flowO2=self.lineEdit_flowO2.text()
        if O2Check==1 and qualO2=="" and phiO2=="" and pressureO2=="" and flowO2=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "O2 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra Arg
        if self.checkBox_Arg.isChecked():
            ArgCheck=1
        else:
            ArgCheck=0
        qualArg=self.lineEdit_qualArg.text()
        phiArg=self.lineEdit_phiArg.text()
        pressureArg=self.lineEdit_pressureArg.text()
        flowArg=self.lineEdit_flowArg.text()
        if ArgCheck==1 and qualArg=="" and phiArg=="" and pressureArg=="" and flowArg=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Arg CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra Oil
        if self.checkBox_Oil.isChecked():
            OilCheck=1
        else:
            OilCheck=0
        qualOil=self.lineEdit_qualOil.text()
        phiOil=self.lineEdit_phiOil.text()
        pressureOil=self.lineEdit_pressureOil.text()
        flowOil=self.lineEdit_flowOil.text()
        if OilCheck==1 and qualOil=="" and phiOil=="" and pressureOil=="" and flowOil=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "Oil CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra khí thải
        if self.checkBox_Emision.isChecked():
            EmisionCheck=1
        else:
            EmisionCheck=0
        qualEmision=self.lineEdit_qualEmision.text()
        phiEmision=self.lineEdit_phiEmision.text()
        pressureEmision=self.lineEdit_pressureEmision.text()
        flowEmision=self.lineEdit_flowEmision.text()
        if EmisionCheck==1 and qualEmision=="" and phiEmision=="" and pressureEmision=="" and flowEmision=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "排気 CheckBoxを選択しましたが何も入力していません！"))
            return
        #kiểm tra nước sạch
        if self.checkBox_CleanWater.isChecked():
            CleanWaterCheck=1
        else:
            CleanWaterCheck=0
        qualCleanWater=self.lineEdit_qualCleanWater.text()
        phiCleanWater=self.lineEdit_phiCleanWater.text()
        pressureCleanWater=self.lineEdit_pressureCleanWater.text()
        flowCleanWater=self.lineEdit_flowCleanWater.text()
        if CleanWaterCheck==1 and qualCleanWater=="" and phiCleanWater=="" and pressureCleanWater=="" and flowCleanWater=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "純水 CheckBoxを選択しましたが何も入力していません！"))
            return
        ###kiểm tra nước làm mát
        if self.checkBox_CoolingWater.isChecked():
            CoolingWaterCheck=1
        else:
            CoolingWaterCheck=0
        qualCoolingWater=self.lineEdit_qualCoolingWater.text()
        phiCoolingWater=self.lineEdit_phiCoolingWater.text()
        pressureCoolingWater=self.lineEdit_pressureCoolingWater.text()
        flowCoolingWater=self.lineEdit_flowCoolingWater.text()
        if CoolingWaterCheck==1 and qualCoolingWater=="" and phiCoolingWater=="" and pressureCoolingWater=="" and flowCoolingWater=="":
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "冷却水 CheckBoxを選択しましたが何も入力していません！"))
            return
        try:
            Photo = self.convertToBinaryData(imagePath)
            InFactoryPhoto = self.convertToBinaryData(FactoryImagePath)
        except:
            self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "十分に写真をアップロードしてください！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "本気ですか？"))
        if respond==QMessageBox.Ok:
            try:
                sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
                cursor = sqliteConnection.cursor()
                sqlite_insert_new_record_query= f"""INSERT INTO Equipment VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,\
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
                data_tuple = (EquipmentName, Model, Serial, DateOfManufacture, Money, Maker, Length, Width, Height, Weight, DateInput, \
                    Invoice, Supporter, EquipmentID, PropertyType, FixedPropertyNumberSharp, ProcessName, Location, FixedPropertyNumberSMV, VolType, \
                    Voltage, Current, RatedPower, ConsumedPower, AirCheck, qualAir, phiAir, pressureAir, flowAir, N2Check, qualN2, phiN2, pressureN2, \
                    flowN2, VACCheck, qualVAC, phiVAC, pressureVAC, flowVAC, O2Check, qualO2, phiO2, pressureO2, flowO2, ArgCheck, qualArg, phiArg, \
                    pressureArg, flowArg, OilCheck, qualOil, phiOil, pressureOil, flowOil, EmisionCheck, qualEmision, phiEmision, pressureEmision, \
                    flowEmision, CleanWaterCheck, qualCleanWater, phiCleanWater, pressureCleanWater, flowCleanWater, CoolingWaterCheck, qualCoolingWater, \
                    phiCoolingWater, pressureCoolingWater, flowCoolingWater, Photo, InFactoryPhoto)
                cursor.execute(sqlite_insert_new_record_query, data_tuple)
                sqliteConnection.commit()
                print("Bản ghi bao gồm các trường dữ liệu và hình ảnh đã được thêm thành công vào database")
                cursor.close()
            except sqlite3.Error as error:
                print("Thêm bản ghi thất bại", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite")
                    self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "新規作成成功！"))
                    PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            #cập nhật logger thông tin thiết bị nhập
            try:
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                newOperate='This user added the equipment named '+EquipmentName+',serial '+Serial+'.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 

        self.ResetFieldForm()
        self.ResetFieldMaintenanceHistory()
        self.UpdateSearchUI()

    def DeleteForm(self):
        if self.GetUserName()=="guest":
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("Equipment", "エラー"),QtWidgets.QApplication.translate("Equipment", "このアカウントはデータを編集する許可がありません！"))
            return
        respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "本気ですか？"))
        if respond==QMessageBox.Ok:
            try:
                sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
                cursor = sqliteConnection.cursor()
                sqlite_delete_record_query= f"""DELETE from Equipment where RowID=?"""
                data_tuple = (key,)
                cursor.execute(sqlite_delete_record_query, data_tuple)
                sqliteConnection.commit()
                print("Đã xóa thành công bản ghi của bảng Equipment")
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite của bảng Equipment")

            try:
                sqliteConnection  = sqlite3.connect('Data.db')
                cursor = sqliteConnection.cursor()
                sqlite_delete_record_query= f"""DELETE from MaintenanceHistory where 製造番号=?"""
                data_tuple = (serial,)
                cursor.execute(sqlite_delete_record_query, data_tuple)
                sqliteConnection.commit()
                print("Đã xóa thành công bản ghi của bảng MaintenanceHistory")
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close()
                    print("Đã đóng kết nối sqlite của bảng MaintenanceHistory")
            self.MessageBoxOK("Information",QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "消去成功！"))
            #cập nhật logger thông tin xóa form thiết bị
            try:
                sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
                cursor = sqliteConnection.cursor()
                query=f"""SELECT *
                FROM UserHistory""" 
                df=pd.read_sql(query,con = sqliteConnection)
                lastRow=len(df)
                oldOperate=df.loc[lastRow-1].iat[3]
                EquipmentName=self.lineEdit_EqName.text()
                Serial=self.lineEdit_serial.text()
                newOperate='This user deleted the equipment named '+EquipmentName+',serial '+Serial+'.'
                if oldOperate=="":
                    combineOperate=newOperate
                else:
                    combineOperate=oldOperate+'\n'+newOperate
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                    SET Session=?
                                    WHERE rowid = ?"""
                data_tuple = (combineOperate,lastRow)
                cursor.execute(sqlite_insert_logger_query, data_tuple)
                sqliteConnection.commit()
                cursor.close()
            except sqlite3.Error as error:
                print("Lỗi", error)
            finally:
                if sqliteConnection:
                    sqliteConnection.close() 
            PyQt5.QtCore.QTimer.singleShot(0, self.SearchText.setFocus)
            ###cập nhật giao diện 
            self.UpdateSearchUI()
            self.ResetFieldForm()
            self.ResetFieldMaintenanceHistory()

    def SaveExcel(self):
        sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
        cursor = sqliteConnection.cursor()
        try:
            query=f"""SELECT *
                FROM Equipment
                WHERE RowID=?""" 
            df=pd.read_sql(query,con = sqliteConnection, params=(key,))
            sqliteConnection.commit()
            cursor.close()

            defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
            des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("Equipment", "Excelを保存するフォルダを選択してください!"),directory=defaultfolder)
            if not des_folderpath=="":
                FileName="/設備情報Exported_On"+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+".xlsx"
                full_path=des_folderpath+FileName
                path=TemplateExcelPath
                ref_workbook=load_workbook(path)
                ws = ref_workbook.active

                List=['D6','M6','W6','D8','M8','W8','D10','J10','P10','W10','D13','M13','W13','D15','M15','W15',\
                    'D17','M17','W17',]
                i=0
                for cell in List:
                    top_left_cell = ws[cell]
                    top_left_cell.value = df.iat[0,i]
                    i=i+1

                if df.iat[0,19]==3:
                    top_left_cell = ws['G21']
                else:
                    top_left_cell = ws['D21']
                top_left_cell.value = df.iat[0,20]

                List=['M20','AA20','AA21','A26','C26','E26','G26','J26','L26','N26','P26','S26','V26','X26','Z26',\
                    'A31','C31','E31','G31','J31','L31','N31','P31','S31','V31','X31','Z31',\
                    'A36','C36','E36','G36','J36','L36','N36','P36','S36','V36','X36','Z36']
                i=21
                for cell in List:
                    if cell=='A26':
                        i=24
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='J26':
                        i=29
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='S26':
                        i=34
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='A31':
                        i=39
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='J31':
                        i=44
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='S31':
                        i=49
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='A36':
                        i=54
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='J36':
                        i=59
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1
                    if cell=='S36':
                        i=64
                        top_left_cell = ws[cell]
                        top_left_cell.value = df.iat[0,i]
                        i=i+1

                    top_left_cell = ws[cell]
                    top_left_cell.value = df.iat[0,i]
                    i=i+1

                equipPhoto=df.iat[0,69]
                global photoPath
                photoPath = application_path + r"\photo.jpg"
                self.writeTofile(equipPhoto, photoPath)
                img=openpyxl.drawing.image.Image(photoPath)
                img.width=20*15
                img.height=14.25*13
                ws.add_image(img,'B40')

                equipPhotoInLine=df.iat[0,70]
                global photoInLinePath
                photoInLinePath = application_path + r"\photoInLine.jpg"
                self.writeTofile(equipPhotoInLine, photoInLinePath)
                img=openpyxl.drawing.image.Image(photoInLinePath)
                img.width=20*20
                img.height=14.25*13
                ws.add_image(img,'N40')

                query=f"""SELECT *
                    FROM MaintenanceHistory
                    WHERE 製造番号=?""" 
                df=pd.read_sql(query,con = sqliteConnection, params=(serial,))
                sqliteConnection.commit()
                cursor.close()
                if df.size==0:   
                    pass
                else:
                    for i in range(0,len(df)):
                        row=str(54+i)
                        top_left_cell = ws['A'+row]
                        top_left_cell.value = df.iat[i,2]
                        top_left_cell = ws['E'+row]
                        top_left_cell.value = df.iat[i,3]
                        top_left_cell = ws['L'+row]
                        top_left_cell.value = df.iat[i,4]
                        top_left_cell = ws['W'+row]
                        top_left_cell.value = df.iat[i,5]
                        top_left_cell = ws['Z'+row]
                        top_left_cell.value = df.iat[i,6]

                ref_workbook.save(full_path)
                #hỏi người dùng có muốn xem file excel
                respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("Equipment", "確認"),QtWidgets.QApplication.translate("Equipment", "Excelファイルが正常に保存されました！\nExcelファイルを開きますか？"))
                if respond==QMessageBox.Ok:
                    os.chdir(des_folderpath)
                    cut_string=FileName.split('/')
                    new_string = cut_string[1]
                    cmd="start excel.exe "+new_string
                    subprocess.check_output(cmd, shell=True)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()
                self.ResetFieldForm()
                self.ResetFieldMaintenanceHistory()

    def UpdateSearchUI(self):
        sqliteConnection  = sqlite3.connect(DatabaseEquipmentPath)
        cursor = sqliteConnection.cursor()
        try:
            sortRowIDQuery=f"""SELECT * FROM Equipment"""
            df=df=pd.read_sql(sortRowIDQuery,con = sqliteConnection)
            df.to_sql('Equipment', con=sqliteConnection, if_exists='replace',index=False)
            
            query=f"""SELECT RowID, 設備名, 製造番号, 設備ID
                    FROM Equipment""" 
            df=pd.read_sql(query,con = sqliteConnection)
            sqliteConnection.commit()
            cursor.close()
            self.model = MaintenanceHistoryTableModel(df)
            self.tableView_Search.setModel(self.model)
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close()

class DashBoard(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(DashBoard,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/DashBoard.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        self.InitialValue()
        self.Signal_Slot()

    def __del__(self):
        print("hàm hủy của DashBoard đã được gọi!")

    def InitialValue(self):
        self.label_2.setText(QtWidgets.QApplication.translate("DashBoard", "続行するには1つ選択してください！"))
        self.label_3.setText(QtWidgets.QApplication.translate("DashBoard", "設備"))
        self.label_4.setText(QtWidgets.QApplication.translate("DashBoard", "在庫"))
        self.label_6.setText(QtWidgets.QApplication.translate("DashBoard", "ベンダー"))
        self.label_5.setText(QtWidgets.QApplication.translate("DashBoard", "データベース"))
        self.label_7.setText(QtWidgets.QApplication.translate("DashBoard", "購入"))
        self.label_8.setText(QtWidgets.QApplication.translate("DashBoard", "ロガー"))
        self.ExitButton.setText(QtWidgets.QApplication.translate("DashBoard", "終了"))
        self.ReturnButton.setText(QtWidgets.QApplication.translate("DashBoard", "戻る"))

        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # update sau mỗi giây
        self.showTime()

    def Signal_Slot(self):
        self.Into_Equipment.clicked.connect(self.IntoEquipmentWindow)
        self.ReturnButton.clicked.connect(self.BackToLogin)
        self.Into_Inventory.clicked.connect(self.IntoInventoryWindow)
        self.Into_Log.clicked.connect(self.IntoLoggerWindow)
        self.Into_Database.clicked.connect(self.IntoDatabaseManagementWindow)
        self.ExitButton.clicked.connect(self.ExitSystem)
        self.shortcut_exit = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_exit.activated.connect(self.ExitSystemQuick)
        self.shortcut_returnToDashBoard = QShortcut(QKeySequence('Ctrl+B'), self)
        self.shortcut_returnToDashBoard.activated.connect(self.BackToLogin)
        self.shortcut_toEquipmentWindow = QShortcut(QKeySequence('Ctrl+E'), self)
        self.shortcut_toEquipmentWindow.activated.connect(self.IntoEquipmentWindow)
        self.shortcut_toInventoryWindow = QShortcut(QKeySequence('Ctrl+I'), self)
        self.shortcut_toInventoryWindow.activated.connect(self.IntoInventoryWindow)
        self.shortcut_toLoggerWindow = QShortcut(QKeySequence('Ctrl+L'), self)
        self.shortcut_toLoggerWindow.activated.connect(self.IntoLoggerWindow)
        self.shortcut_toIntoDatabaseManagementWindow = QShortcut(QKeySequence('Ctrl+D'), self)
        self.shortcut_toIntoDatabaseManagementWindow.activated.connect(self.IntoDatabaseManagementWindow)
        
    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_clock.setText(displayTxt)

    def IntoInventoryWindow(self):
        inv=Inventory()
        widget.addWidget(inv)
        widget.setFixedHeight(668)
        widget.setFixedWidth(1173)
        widget.setWindowTitle("Inventory Window")
        widget.setWindowIcon(QIcon(":icon/sys.png"))
        #widget.removeWidget(widget.currentWidget())
        widget.currentWidget().deleteLater()
        widget.setCurrentWidget(inv)
        widget.move(screen_center(widget))
        
    def IntoEquipmentWindow(self):
        Equip=Equipment()
        widget.addWidget(Equip)
        widget.setFixedHeight(721)
        widget.setFixedWidth(1301)
        widget.setWindowTitle("Equipment Window")
        widget.setWindowIcon(QIcon(":icon/Eq.png"))
        #widget.removeWidget(widget.currentWidget())
        widget.currentWidget().deleteLater()
        widget.setCurrentWidget(Equip)
        widget.move(screen_center(widget))

    def IntoLoggerWindow(self):
        log=Logger()
        widget.addWidget(log)
        widget.setFixedHeight(604)
        widget.setFixedWidth(800)
        widget.setWindowTitle("Logger Window")
        widget.setWindowIcon(QIcon(":icon/historyUser.png"))
        #widget.removeWidget(widget.currentWidget())
        widget.currentWidget().deleteLater()
        widget.setCurrentWidget(log)
        widget.move(screen_center(widget))
    
    def IntoDatabaseManagementWindow(self):
        dm=DatabaseManagement()
        widget.addWidget(dm)
        widget.setFixedHeight(604)
        widget.setFixedWidth(800)
        widget.setWindowTitle("Database Management Window")
        widget.setWindowIcon(QIcon(":icon/DatabaseIcon.png"))
        #widget.removeWidget(widget.currentWidget())
        widget.currentWidget().deleteLater()
        widget.setCurrentWidget(dm)
        widget.move(screen_center(widget))

    def BackToLogin(self):
        #xóa database tạm của Inventory Window
        print("remove language")
        try:
            sqliteConnection  = sqlite3.connect(DatabaseInventoryPath)
            cursor      = sqliteConnection.cursor()
            dropInventoryTable = "DROP TABLE IF EXISTS Inventory"
            cursor.execute(dropInventoryTable)
            dropsearchResultTable = "DROP TABLE IF EXISTS SearchResult"
            cursor.execute(dropsearchResultTable)
            dropsearchResultTable = "DROP TABLE IF EXISTS Filter"
            cursor.execute(dropsearchResultTable)
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 
        #cập nhật file log thời gian đăng xuất
        try:
            sqliteConnection  = sqlite3.connect(DatabaseLoggerPath)
            cursor = sqliteConnection.cursor()
            query=f"""SELECT Session
            FROM UserHistory""" 
            df=pd.read_sql(query,con = sqliteConnection)
            lastRow=len(df)
            if df.loc[lastRow-1].iat[0]=="":
                noOperation='This user read only, the database is not changed.'
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                SET "Logout Time"=?,
                                    Session=?
                                WHERE rowid = ?"""
                data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),noOperation,lastRow)
            else:
                sqlite_insert_logger_query= f"""UPDATE UserHistory 
                                SET "Logout Time"=?
                                WHERE rowid = ?"""
                data_tuple = (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),lastRow)
            cursor.execute(sqlite_insert_logger_query, data_tuple)
            sqliteConnection.commit()
            cursor.close()
        except sqlite3.Error as error:
            print("Lỗi", error)
        finally:
            if sqliteConnection:
                sqliteConnection.close() 

        widget.setFixedHeight(604)
        widget.setFixedWidth(800)
        widget.setWindowTitle("Login Window")
        widget.setWindowIcon(QIcon(":icon/LoginIcon.png"))
        widget.currentWidget().deleteLater()
        widget.setCurrentIndex(widget.currentIndex()-1)
        widget.move(screen_center(widget))

class DatabaseManagement(QtWidgets.QMainWindow,ShareTheSameMethod):
    def __init__(self):
        super(DatabaseManagement,self).__init__()
        self.setAttribute(PyQt5.QtCore.Qt.WA_DeleteOnClose, True)
        fileh = PyQt5.QtCore.QFile(':/ui/DatabaseManagement.ui')
        fileh.open(PyQt5.QtCore.QFile.ReadOnly)
        uic.loadUi(fileh, self)
        fileh.close()
        self.InitialValue()
        self.Signal_Slot()

    def __del__(self):
        print("hàm hủy của DatabaseManagement đã được gọi!")

    def InitialValue(self):
        self.label_2.setText(QtWidgets.QApplication.translate("DatabaseManagement", "データベース管理"))
        self.ExitButton.setText(QtWidgets.QApplication.translate("DatabaseManagement", "終了"))
        self.BackupButton.setText(QtWidgets.QApplication.translate("DatabaseManagement", "保存"))
        self.RestoreButton.setText(QtWidgets.QApplication.translate("DatabaseManagement", "復元"))
        self.DeleteButton.setText(QtWidgets.QApplication.translate("DatabaseManagement", "削除"))
        self.ReturnButton.setText(QtWidgets.QApplication.translate("DatabaseManagement", "戻る"))
        
        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000) # update sau mỗi giây
        self.showTime()
        global listDatabase
        listDatabase=[]
        listStringDatabase=[]
        if os.path.exists(DatabaseEquipmentPath):
            listDatabase.append(DatabaseEquipmentPath)
            listStringDatabase.append("Equipment Database")
        if os.path.exists(DatabaseInventoryPath):
            listDatabase.append(DatabaseInventoryPath)
            listStringDatabase.append("Inventory Database")
        if os.path.exists(DatabaseLoggerPath):
            listDatabase.append(DatabaseLoggerPath)
            listStringDatabase.append("Logger Database")
        self.comboBox.addItems(listStringDatabase)
        self.ShowAllInfor()

    def ShowAllInfor(self):
        dataframe=pd.DataFrame(columns = ["File Name","Created Date","Modified Date","File Size In Mb"])
        print(listDatabase)
        for x in listDatabase:
            listitem=[]
            filename = os.path.basename(x)
            listitem.append(filename)
            dateCreate=time.strftime("%Y/%m/%d %H:%M:%S",time.strptime(time.ctime(os.path.getctime(x))))
            listitem.append(dateCreate)
            dateMod=time.strftime("%Y/%m/%d %H:%M:%S",time.strptime(time.ctime(os.path.getmtime(x))))
            listitem.append(dateMod)
            fileSize=os.stat(x).st_size/(1024 * 1024)
            listitem.append(fileSize)
            df_length = len(dataframe)
            dataframe.loc[df_length] = listitem
        print(dataframe)
        self.model = PandasModelForExportHistory(dataframe)
        self.tableView_Database.setModel(self.model)

    def showTime(self):
        currentTime = QDateTime.currentDateTime()
        displayTxt = currentTime.toString('yyyy/MM/dd hh:mm:ss')
        self.label_clock.setText(displayTxt)
    
    def Signal_Slot(self):
        self.shortcut_exit = QShortcut(QKeySequence('Ctrl+S'), self)
        self.shortcut_exit.activated.connect(self.ExitSystemQuick)
        self.shortcut_returnToDashBoard = QShortcut(QKeySequence('Ctrl+B'), self)
        self.shortcut_returnToDashBoard.activated.connect(self.BackToDashBoard)

        self.ReturnButton.clicked.connect(self.BackToDashBoard)
        self.ExitButton.clicked.connect(self.ExitSystem)
        self.BackupButton.clicked.connect(self.BackupDatabase)
        self.DeleteButton.clicked.connect(self.DeleteDatabase)
        self.RestoreButton.clicked.connect(self.RestoreDatabase)
    
    def BackupDatabase(self):
        if self.GetUserName()=="admin":
            defaultfolder='C:\\Users\\{0}\\Documents'.format(os.getlogin())
            des_folderpath = QtWidgets.QFileDialog.getExistingDirectory(None, caption=QtWidgets.QApplication.translate("DatabaseManagement", "フォルダを選択してください!"),directory=defaultfolder)
            if not des_folderpath=="":
                selectedDatabase=str(self.comboBox.currentText())
                if "Inventory" in selectedDatabase:
                    shutil.copyfile(DatabaseInventoryPath, des_folderpath+r'\InventoryDatabase.db')
                    shutil.copyfile(DatabaseInventoryPath, application_path+r'\Backup Database\InventoryDatabase.db')
                    #os.rename(des_folderpath+r'\InventoryDatabase.db',des_folderpath+r'\InventoryDatabaseBackupIn'+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+'.db')
                elif "Equipment" in selectedDatabase:
                    shutil.copyfile(DatabaseEquipmentPath, des_folderpath+r'\EquipmentDatabase.db')
                    shutil.copyfile(DatabaseEquipmentPath, application_path+r'\Backup Database\EquipmentDatabase.db')
                    #os.rename(des_folderpath+r'\EquipmentDatabase.db',des_folderpath+r'\EquipmentDatabaseBackupIn'+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+'.db')
                elif "Logger" in selectedDatabase:
                    shutil.copyfile(DatabaseLoggerPath, des_folderpath+r'\LoggerDatabase.db')
                    shutil.copyfile(DatabaseLoggerPath, application_path+r'\Backup Database\LoggerDatabase.db')
                    #os.rename(des_folderpath+r'\LoggerDatabase.db',des_folderpath+r'\LoggerDatabaseBackupIn'+datetime.now().strftime("%d_%m_%YAt%H_%M_%S")+'.db')
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "データベースを正常にバックアップしました！"))
        else:
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("DatabaseManagement", "エラー"),QtWidgets.QApplication.translate("DatabaseManagement", "この機能には管理者が必要です！"))
    
    def DeleteDatabase(self):
        if self.GetUserName()=="admin":
            respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "本気ですか？"))
            if respond==QMessageBox.Ok:
                selectedDatabase=str(self.comboBox.currentText())
                if "Inventory" in selectedDatabase:
                    conn = sqlite3.connect(DatabaseInventoryPath)
                    cursor = conn.cursor()
                    cursor.execute('DELETE FROM ImportHistory;',)
                    cursor.execute('DELETE FROM ExportHistory;',)
                    conn.commit()
                    conn.close()
                    self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Inventory Databaseを正常に削除しました！"))

                elif "Equipment" in selectedDatabase:
                    conn = sqlite3.connect(DatabaseEquipmentPath)
                    cursor = conn.cursor()
                    cursor.execute('DELETE FROM Equipment;',)
                    cursor.execute('DELETE FROM MaintenanceHistory;',)
                    conn.commit()
                    conn.close()
                    self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Equipment Databaseを正常に削除しました！"))

                elif "Logger" in selectedDatabase:
                    respond=self.MessageBoxOKCancel(QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Logger Databaseを削除するとシステムが再起動します。続行しますか？"))
                    if respond==QMessageBox.Ok:
                        conn = sqlite3.connect(DatabaseLoggerPath)
                        cursor = conn.cursor()
                        cursor.execute('DELETE FROM UserHistory;',)	
                        conn.commit()
                        conn.close()
                        self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Logger Databaseを正常に削除しました！"))
                        PyQt5.QtCore.QCoreApplication.quit()
                        PyQt5.QtCore.QProcess.startDetached(sys.executable, sys.argv)
        else:
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("DatabaseManagement", "エラー"),QtWidgets.QApplication.translate("DatabaseManagement", "この機能には管理者が必要です！"))
    
    def RestoreDatabase(self):
        if self.GetUserName()=="admin":
            InventoryIsBackuped=os.path.exists(application_path+r'\Backup Database\InventoryDatabase.db')
            EquipmentIsBackuped=os.path.exists(application_path+r'\Backup Database\EquipmentDatabase.db')
            LoggerIsBackuped=os.path.exists(application_path+r'\Backup Database\LoggerDatabase.db')
            if not InventoryIsBackuped and not EquipmentIsBackuped and not LoggerIsBackuped:
                self.MessageBoxOK("Critical",QtWidgets.QApplication.translate("DatabaseManagement", "エラー"),QtWidgets.QApplication.translate("DatabaseManagement", "バックアップデータベースはありません！"))           
                return
            if InventoryIsBackuped:
                shutil.copyfile(application_path+r'\Backup Database\InventoryDatabase.db', DatabaseInventoryPath)
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Inventory Databaseを正常に復元しました！"))
            if EquipmentIsBackuped:
                shutil.copyfile(application_path+r'\Backup Database\EquipmentDatabase.db', DatabaseEquipmentPath)
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Equipment Databaseを正常に復元しました！"))
            if LoggerIsBackuped:
                shutil.copyfile(application_path+r'\Backup Database\LoggerDatabase.db', DatabaseLoggerPath) 
                self.MessageBoxOK("Information",QtWidgets.QApplication.translate("DatabaseManagement", "確認"),QtWidgets.QApplication.translate("DatabaseManagement", "Logger Databaseを正常に復元しました！"))
        else:
            self.MessageBoxOK("Warning",QtWidgets.QApplication.translate("DatabaseManagement", "エラー"),QtWidgets.QApplication.translate("DatabaseManagement", "この機能には管理者が必要です！"))
#############################################CHƯƠNG TRÌNH CHÍNH MAIN########################################

if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
elif __file__:
    application_path = os.path.dirname(__file__)
DatabaseEquipmentPath=application_path+r"\EquipmentDatabase.db"
DatabaseInventoryPath=application_path+r"\InventoryDatabase.db"
DatabaseLoggerPath=application_path+r"\LoggerDatabase.db"
DefaultPhotoPath=application_path+r"\addImage.png"
TemplateExcelPath=application_path+r"\Template.xlsx"
screen_center = lambda widget: QApplication.desktop().screen().rect().center()- widget.rect().center()
TodayIs=datetime.today().strftime("%Y/%m/%d")

app=QApplication(sys.argv)
app.setWindowIcon(PyQt5.QtGui.QIcon(':icon/window.png'))
widget=QtWidgets.QStackedWidget()
widget.setWindowFlag(PyQt5.QtCore.Qt.WindowCloseButtonHint, False)
LoginSys=Login()
widget.addWidget(LoginSys)
widget.setWindowTitle("Login Window")
widget.setWindowIcon(PyQt5.QtGui.QIcon(':icon/LoginIcon.png'))
widget.setFixedHeight(604)
widget.setFixedWidth(800)
widget.show()
app.beep()
app.aboutQt()
try:
    sys.exit(app.exec_())
except:
    print("existing")




